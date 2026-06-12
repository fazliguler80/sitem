# bina/admin.py

import os
import zipfile
import tempfile
import shutil
import math
import json
from datetime import datetime, timedelta
from decimal import Decimal

from django import forms
from django.contrib import admin
from django.contrib.admin import SimpleListFilter
from django.contrib.admin.views.main import ChangeList
from django.contrib.admin.options import IncorrectLookupParameters
from django.contrib.auth.models import User
from django.contrib.auth import admin as auth_admin
from django.contrib import messages
from django.contrib.auth.forms import PasswordChangeForm
from django.urls import path
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
from django.contrib.admin import SimpleListFilter
from datetime import date

from .models import (
    SiteAyarlari, Blok, Daire, Kisi, DaireIliskisi, 
    Aidat, Gider, Yonetici, Personel, Abonelik, 
    Firma, Banka, BankaHareket, Depozito, DepozitoHareket,
    Fatura, DaireKullanici, AsgariUcret, MaasBordrosu
)

# bina/admin.py - EN ÜSTE EKLEYİN (import'lardan sonra)
from django import forms
from django.shortcuts import render
from django.urls import path
from django.utils.html import format_html
from datetime import datetime, timedelta

class TarihFiltresiMixin:
    """Admin liste ekranlarına hızlı tarih filtresi ekleyen mixin"""
    
    def changelist_view(self, request, extra_context=None):
        # Hızlı tarih filtrelerini işle
        hizli_tarih = request.GET.get('hizli_tarih', '')
        baslangic_tarih = request.GET.get('baslangic_tarih', '')
        bitis_tarih = request.GET.get('bitis_tarih', '')
        
        bugun = datetime.now().date()
        
        if hizli_tarih == 'son7gun':
            baslangic = bugun - timedelta(days=7)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son15gun':
            baslangic = bugun - timedelta(days=15)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son30gun':
            baslangic = bugun - timedelta(days=30)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son1ay':
            baslangic = bugun - timedelta(days=30)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son3ay':
            baslangic = bugun - timedelta(days=90)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son6ay':
            baslangic = bugun - timedelta(days=180)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son1yil':
            baslangic = bugun - timedelta(days=365)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        
        # Tarih aralığını queryset'e uygula
        if baslangic_tarih and bitis_tarih:
            try:
                baslangic = datetime.strptime(baslangic_tarih, '%Y-%m-%d').date()
                bitis = datetime.strptime(bitis_tarih, '%Y-%m-%d').date()
                request.GET._mutable = True
                # Tarih aralığını URL parametresi olarak ekle
                if hasattr(self, 'tarih_alan'):
                    request.GET[f'{self.tarih_alan}__range'] = f'{baslangic},{bitis}'
                request.GET._mutable = False
            except:
                pass
        
        extra_context = extra_context or {}
        extra_context['hizli_tarih'] = hizli_tarih
        extra_context['baslangic_tarih'] = baslangic_tarih
        extra_context['bitis_tarih'] = bitis_tarih
        
        return super().changelist_view(request, extra_context=extra_context)
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('tarih-filtrele/', self.admin_site.admin_view(self.tarih_filtrele_view), name=f'{self.model._meta.model_name}_tarih_filtrele'),
        ]
        return custom_urls + urls
    
    def tarih_filtrele_view(self, request):
        """AJAX ile tarih filtresi uygulama"""
        from django.http import JsonResponse
        hizli_tarih = request.GET.get('hizli_tarih', '')
        baslangic = request.GET.get('baslangic', '')
        bitis = request.GET.get('bitis', '')
        
        # URL'yi oluştur
        url = f'../?'
        if hizli_tarih:
            url += f'hizli_tarih={hizli_tarih}&'
        if baslangic:
            url += f'baslangic_tarih={baslangic}&'
        if bitis:
            url += f'bitis_tarih={bitis}&'
        
        return JsonResponse({'url': url})

# ==================== BLOK FİLTRESİ ====================
class BlokFilter(SimpleListFilter):
    """Daire ilişkileri için blok bazlı filtreleme"""
    title = 'Blok'
    parameter_name = 'blok'
    
    def lookups(self, request, model_admin):
        # Tüm blokları listele
        bloklar = Blok.objects.all().order_by('blok_adi')
        return [(blok.blok_adi, f'{blok.get_blok_adi_display()}') for blok in bloklar]
    
    def queryset(self, request, queryset):
        if self.value():
            # Seçilen bloktaki daireleri filtrele
            return queryset.filter(daire__blok__blok_adi=self.value())
        return queryset


# ==================== DAİRE İLİŞKİLERİ ADMIN ====================
class DaireIliskisiAdmin(admin.ModelAdmin):
    list_display = ['daire_bilgisi', 'kisi_bilgisi', 'iliski_tipi', 'baslangic_tarihi', 'bitis_tarihi', 'aktif_mi', 'kira_tutari']
    list_filter = ['aktif_mi', 'iliski_tipi', BlokFilter]
    list_editable = ['aktif_mi', 'birincil_mi']
    search_fields = ['daire__daire_no', 'daire__blok__blok_adi', 'kisi__ad_soyad', 'kisi__telefon']
    list_editable = ['aktif_mi']
    list_per_page = 20
    
    fieldsets = (
        ('Daire Bilgileri', {
            'fields': ('daire',)
        }),
        ('Kişi Bilgileri', {
            'fields': ('kisi', 'iliski_tipi', 'baslangic_tarihi', 'bitis_tarihi', 'aktif_mi')
        }),
        ('Önemli Ayarlar', {
            'fields': ('birincil_mi', 'kira_tutari'),  # SADECE BURADA BİR KERE
            'description': '✅ Birincil kişi: Aidat ve borçlar bu kişiye yansır. Her dairede sadece 1 kişi birincil olabilir.'
        }),
    )
    
    def daire_bilgisi(self, obj):
        """Daire bilgisini formatlı göster"""
        blok_adi = obj.daire.blok.get_blok_adi_display()
        return format_html(
            '<strong>{}</strong><br><small style="color: #666;">{}</small>',
            f"{blok_adi} - {obj.daire.daire_no}",
            f"Arsa Payı: {obj.daire.arsa_pay_pay}/{obj.daire.arsa_pay_payda}"
        )
    daire_bilgisi.short_description = 'Daire'
    
    def kisi_bilgisi(self, obj):
        """Kişi bilgisini formatlı göster"""
        return format_html(
            '<strong>{}</strong><br><small>📞 {}</small>',
            obj.kisi.ad_soyad,
            obj.kisi.telefon
        )
    kisi_bilgisi.short_description = 'Kişi'
    
    actions = ['aktif_yap', 'pasif_yap']
    
    def aktif_yap(self, request, queryset):
        queryset.update(aktif_mi=True)
        self.message_user(request, f"{queryset.count()} ilişki aktif yapıldı.")
    aktif_yap.short_description = "Seçili ilişkileri aktif yap"
    
    def pasif_yap(self, request, queryset):
        queryset.update(aktif_mi=False)
        self.message_user(request, f"{queryset.count()} ilişki pasif yapıldı.")
    pasif_yap.short_description = "Seçili ilişkileri pasif yap"

class YoneticiAdmin(admin.ModelAdmin):
    list_display = ('ad_soyad', 'gorev_tipi', 'telefon', 'email', 'gorev_baslangic', 'aktif_mi')
    list_filter = ('gorev_tipi', 'aktif_mi')
    search_fields = ('ad_soyad', 'email')

# bina/admin.py - Düzeltilmiş DaireIliskisiInline

class DaireIliskisiInline(admin.TabularInline):
    """Kişi detay sayfasında daire ilişkilerini gösterir"""
    model = DaireIliskisi
    extra = 1
    fields = ['daire', 'iliski_tipi', 'birincil_mi', 'aktif_mi', 'baslangic_tarihi', 'bitis_tarihi']
    raw_id_fields = ['daire']
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "daire":
            kwargs["queryset"] = Daire.objects.select_related('blok').order_by('blok__blok_adi', 'daire_no')
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


class DaireIliskisiInline(admin.TabularInline):
    """Kişi detay sayfasında daire ilişkilerini gösterir"""
    model = DaireIliskisi
    extra = 1
    fields = ['daire', 'iliski_tipi', 'birincil_mi', 'aktif_mi', 'baslangic_tarihi', 'bitis_tarihi']
    raw_id_fields = ['daire']  # Autocomplete yerine raw_id_fields kullan
    # autocomplete_fields = ['daire']  # Bunu yorum satırı yapın veya silin
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "daire":
            kwargs["queryset"] = Daire.objects.select_related('blok').order_by('blok__blok_adi', 'daire_no')
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


class DaireIliskisiInlineForDaire(admin.TabularInline):
    """Daire detay sayfasında kişi ilişkilerini gösterir"""
    model = DaireIliskisi
    extra = 1
    fields = ['kisi', 'iliski_tipi', 'birincil_mi', 'aktif_mi', 'baslangic_tarihi', 'bitis_tarihi']
    raw_id_fields = ['kisi']
    autocomplete_fields = ['kisi']
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "kisi":
            kwargs["queryset"] = Kisi.objects.order_by('ad_soyad')
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

class KisiAdmin(admin.ModelAdmin):
    list_display = ['ad_soyad', 'kisi_tipi', 'diger_aciklama', 'telefon', 'email', 'aktif_mi', 'daire_bilgisi']
    list_filter = ['kisi_tipi', 'aktif_mi']
    search_fields = ['ad_soyad', 'telefon', 'email', 'tc_kimlik', 'diger_aciklama']
    list_editable = ['aktif_mi']
    
    # YENİ: Daire ilişkilerini göster
    inlines = [DaireIliskisiInline]
    
    fieldsets = (
        ('Kişisel Bilgiler', {
            'fields': ('ad_soyad', 'kisi_tipi', 'diger_aciklama', 'tc_kimlik')
        }),
        ('İletişim Bilgileri', {
            'fields': ('telefon', 'cep_telefonu', 'is_telefonu', 'email', 'adres')
        }),
        ('Acil Durum', {
            'fields': ('acil_durum_kisisi', 'acil_durum_tel'),
            'classes': ('collapse',)
        }),
        ('Diğer', {
            'fields': ('aktif_mi', 'notlar'),
            'classes': ('collapse',)
        }),
    )
    
    # YENİ: Listedeki her kişi için bağlı olduğu daireyi göster
    def daire_bilgisi(self, obj):
        iliskiler = obj.daire_iliskileri.filter(aktif_mi=True)
        if iliskiler.exists():
            daireler = []
            for iliski in iliskiler:
                birincil = " (Birincil)" if iliski.birincil_mi else ""
                daireler.append(f"{iliski.daire.blok.get_blok_adi_display()}-{iliski.daire.daire_no}{birincil}")
            return ", ".join(daireler)
        return "-"
    daire_bilgisi.short_description = 'Bağlı Daire(ler)'
    
    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset.prefetch_related('daire_iliskileri__daire__blok')

    class Media:
        js = ('admin/js/kisi_tipi.js',)
        css = {
            'all': ('admin/css/kisi_admin.css',)
        }

class FirmaAdmin(admin.ModelAdmin):
    list_display = ('firma_adi', 'tip', 'yetkili_kisi', 'telefon', 'sozlesme_bitis', 'aktif_mi')
    list_filter = ('tip', 'aktif_mi')
    search_fields = ('firma_adi', 'yetkili_kisi')

class DaireAdmin(admin.ModelAdmin):
    list_display = ('blok', 'daire_no', 'kat', 'daire_tipi', 'malik_bilgisi', 
                    'isletme_giderlerinden_muaf', 'demirbas_giderlerinden_muaf', 'durum')
    list_filter = ('blok', 'daire_tipi', 'isletme_giderlerinden_muaf', 'demirbas_giderlerinden_muaf')
    list_editable = ('isletme_giderlerinden_muaf', 'demirbas_giderlerinden_muaf')
    search_fields = ('daire_no', 'blok__blok_adi')

    def malik_bilgisi(self, obj):
        iliski = obj.iliskiler.filter(aktif_mi=True, iliski_tipi='ev_sahibi').first()
        return iliski.kisi.ad_soyad if iliski and iliski.kisi else '-'
    malik_bilgisi.short_description = 'Malik'

    def telefon_bilgisi(self, obj):
        iliski = obj.iliskiler.filter(aktif_mi=True, iliski_tipi='ev_sahibi').first()
        return iliski.kisi.telefon if iliski and iliski.kisi else '-'
    telefon_bilgisi.short_description = 'Telefon'

    def durum(self, obj):
        # Boolean değer döndür (True=Dolu, False=Boş)
        return obj.iliskiler.filter(aktif_mi=True).exists()
    durum.boolean = True
    durum.short_description = 'Durum'

class BankaAdmin(admin.ModelAdmin):
    list_display = ('banka_adi', 'hesap_adi', 'iban', 'guncel_bakiye')
    list_filter = ('hesap_tipi', 'aktif_mi')
    search_fields = ('banka_adi', 'hesap_adi', 'iban')

    def save_model(self, request, obj, form, change):
        # Yeni kayıt ise güncel bakiyeyi açılış bakiyesi ile doldur
        if not change:  # Yani yeni ekleniyorsa
            obj.guncel_bakiye = obj.acilis_bakiyesi
        super().save_model(request, obj, form, change)

    def changelist_view(self, request, extra_context=None):
        queryset = self.get_queryset(request)
        # Uygulanan filtreleri al (isteğe bağlı)
        try:
            from django.contrib.admin.views.main import ChangeList
            cl = ChangeList(request, self.model, self.list_display, self.list_display_links,
                            self.list_filter, self.date_hierarchy, self.search_fields,
                            self.list_select_related, self.list_per_page, self.list_max_show_all,
                            self.list_editable, self, self.sortable_by)
            queryset = cl.get_queryset(request)
        except:
            pass
        from django.db.models import Sum
        toplam = queryset.aggregate(Sum('guncel_bakiye'))['guncel_bakiye__sum'] or 0
        extra_context = extra_context or {}
        extra_context['toplam_bakiye'] = toplam
        return super().changelist_view(request, extra_context=extra_context)

class BankaHareketAdmin(TarihFiltresiMixin, admin.ModelAdmin):
    change_list_template = 'admin/banka_hareketleri_filtreli.html'
    tarih_alan = 'tarih'
    list_display = ('banka', 'hareket_tipi', 'tutar', 'tarih', 'aciklama')
    list_filter = ('hareket_tipi', 'tarih', 'banka')
    search_fields = ('aciklama', 'dekont_no')
    date_hierarchy = 'tarih'
    list_per_page = 25
    
    def changelist_view(self, request, extra_context=None):
        # Tarih aralığı filtrelemesi
        tarih_baslangic = request.GET.get('tarih_baslangic', '')
        tarih_bitis = request.GET.get('tarih_bitis', '')
        
        # Filtrelenmiş queryset'i al
        queryset = self.get_queryset(request)
        
        # Tarih aralığı kontrolü
        if tarih_baslangic and tarih_bitis:
            try:
                from datetime import datetime
                baslangic = datetime.strptime(tarih_baslangic, '%Y-%m-%d').date()
                bitis = datetime.strptime(tarih_bitis, '%Y-%m-%d').date()
                if baslangic == bitis:
                    queryset = queryset.filter(tarih=baslangic)
                else:
                    queryset = queryset.filter(tarih__range=[baslangic, bitis])
            except:
                pass
        elif tarih_baslangic:
            try:
                from datetime import datetime
                baslangic = datetime.strptime(tarih_baslangic, '%Y-%m-%d').date()
                queryset = queryset.filter(tarih__gte=baslangic)
            except:
                pass
        elif tarih_bitis:
            try:
                from datetime import datetime
                bitis = datetime.strptime(tarih_bitis, '%Y-%m-%d').date()
                queryset = queryset.filter(tarih__lte=bitis)
            except:
                pass
        
        # Toplam hesaplamaları
        from django.db.models import Sum
        toplam_gelir = queryset.filter(hareket_tipi='gelir').aggregate(Sum('tutar'))['tutar__sum'] or 0
        toplam_gider = queryset.filter(hareket_tipi='gider').aggregate(Sum('tutar'))['tutar__sum'] or 0
        net_bakiye = float(toplam_gelir) - float(toplam_gider)
        
        # Banka listesi (modal için)
        from .models import Banka
        bankalar = Banka.objects.all()
        
        extra_context = extra_context or {}
        extra_context['toplam_gelir'] = toplam_gelir
        extra_context['toplam_gider'] = toplam_gider
        extra_context['net_bakiye'] = net_bakiye
        extra_context['bankalar'] = bankalar
        
        return super().changelist_view(request, extra_context=extra_context)
    
    class Media:
        css = {
            'all': ('admin/css/banka_hareketleri.css',)
        }

# bina/admin.py - AidatAdmin sınıfını tamamen düzeltin

from django.contrib import admin
from django.urls import path
from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse
from django.utils.html import format_html
from django.contrib import messages
from django.contrib.admin.views.main import ChangeList
from django.contrib.admin.options import IncorrectLookupParameters
from django.db.models import Sum
import json

from .models import Aidat, Kisi

# bina/admin.py - BASİT AİDAT ADMIN (geçici çözüm)

# bina/admin.py


# bina/admin.py - AidatAdmin sınıfı

class OdemeDurumuFilter(SimpleListFilter):
    title = 'Ödeme Durumu'
    parameter_name = 'odeme_durumu'
    
    def lookups(self, request, model_admin):
        return (
            ('odendi', '✅ Ödendi'),
            ('odenmedi', '❌ Ödenmedi'),
        )
    
    def queryset(self, request, queryset):
        if self.value() == 'odendi':
            return queryset.filter(odeme_yapildi_mi=True)
        if self.value() == 'odenmedi':
            return queryset.filter(odeme_yapildi_mi=False)
        return queryset

from django.contrib.admin import SimpleListFilter
from bina.models import Daire

class DaireFiltresi(SimpleListFilter):
    title = 'Daire'
    parameter_name = 'daire'
    
    def lookups(self, request, model_admin):
        blok_id = request.GET.get('daire__blok__id__exact', None)
        
        if blok_id:
            daireler = Daire.objects.filter(blok_id=blok_id).order_by('daire_no')
        else:
            daireler = Daire.objects.none()
        
        result = []
        for d in daireler:
            blok_adi = d.blok.get_blok_adi_display()
            arsa_pay = f"{d.arsa_pay_pay}/{d.arsa_pay_payda}"
            # Örnek: "C-18 (48/2700)"
            display_text = f"{blok_adi}-{d.daire_no} ({arsa_pay})"
            result.append((d.id, display_text))
        
        return result
    
    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(daire__id=self.value())
        return queryset


class AidatAdmin(TarihFiltresiMixin, admin.ModelAdmin):
    tarih_alan = 'odeme_tarihi'
    list_display = ('daire', 'ay', 'yil', 'aidat_tipi', 'tutar', 'odeme_yapildi_mi', 'odeme_tarihi', 'kim_odedi_bilgisi')
    list_filter = (
        'odeme_yapildi_mi',
        'yil',
        'ay',
        'aidat_tipi',
        'daire__blok',  # Önce Blok
        DaireFiltresi,             # Sonra Daire (bloğa bağlı)
        'gider__tip',
    )
    search_fields = ('daire__blok__blok_adi', 'daire__daire_no', 'aciklama', 'kim_odedi__ad_soyad')
    date_hierarchy = 'odeme_tarihi'
    list_per_page = 25
    list_editable = ('odeme_yapildi_mi',)
    autocomplete_fields = ['daire']
    
    fieldsets = (
        ('Temel Bilgiler', {
            'fields': ('daire', 'ay', 'yil', 'aidat_tipi', 'tutar', 'aciklama')
        }),
        ('Gider Bilgisi', {
            'fields': ('gider', 'yuvarlama_farki'),
            'classes': ('collapse',)
        }),
        ('Ödeme Bilgileri', {
            'fields': ('odeme_yapildi_mi', 'odeme_tarihi', 'kim_odedi', 'odeme_notu'),
        }),
    )
    
    def save_model(self, request, obj, form, change):
        """Aidat kaydedilirken otomatik açıklama oluştur"""
        
        if obj.aidat_tipi == 'sabit':
            # Sabit aidat: gider yok, açıklama formatlı
            obj.gider = None
            obj.aciklama = f"{obj.ay}/{obj.yil} Sabit Aidat - {obj.tutar:.2f} TL"
        else:
            # Ekstra veya Yakıt aidatları
            if not obj.aciklama or change == False:
                if obj.gider:
                    hesaplanan = float(obj.tutar) - float(obj.yuvarlama_farki)
                    obj.aciklama = f"{obj.gider.get_tip_display()} - {obj.gider.tarih} (Hesap: {hesaplanan:.2f} TL → Ödenecek: {obj.tutar:.2f} TL)"
                else:
                    obj.aciklama = f"{obj.ay}/{obj.yil} {obj.get_aidat_tipi_display()} - {obj.tutar:.2f} TL"
        
        super().save_model(request, obj, form, change)
    
    def kim_odedi_bilgisi(self, obj):
        if obj.kim_odedi:
            return obj.kim_odedi.ad_soyad
        return "-"
    kim_odedi_bilgisi.short_description = "Ödeyen Kişi"

    def gider_bilgisi(self, obj):
        """Gider bilgisini formatlı göster"""
        if obj.gider:
            return f"{obj.gider.get_tip_display()} - {obj.gider.tarih.strftime('%d/%m/%Y')}"
        return "-"
    
    def kim_odedi_bilgisi(self, obj):
        """Ödeyen kişi bilgisini göster"""
        if obj.kim_odedi:
            return f"{obj.kim_odedi.ad_soyad}"
        return "-"
    kim_odedi_bilgisi.short_description = "Ödeyen Kişi"
    kim_odedi_bilgisi.admin_order_field = 'kim_odedi__ad_soyad'
    
    actions = ['toplu_odeme_yap', 'toplu_odeme_iptal', 'aidat_raporu_excel']
    
    def toplu_odeme_yap(self, request, queryset):
        """Seçili aidatları toplu ödeme yap"""
        from datetime import date
        count = 0
        for aidat in queryset:
            if not aidat.odeme_yapildi_mi:
                aidat.odeme_yap(odeme_tarihi=date.today())
                count += 1
        self.message_user(request, f"✅ {count} aidat ödendi.")
    toplu_odeme_yap.short_description = "Seçili aidatları öde"
    
    def toplu_odeme_iptal(self, request, queryset):
        """Seçili aidatların ödemesini iptal et"""
        count = 0
        for aidat in queryset:
            if aidat.odeme_yapildi_mi:
                aidat.odeme_iptal()
                count += 1
        self.message_user(request, f"✅ {count} aidatın ödemesi iptal edildi.")
    toplu_odeme_iptal.short_description = "Seçili aidatların ödemesini iptal et"
    
    def aidat_raporu_excel(self, request, queryset):
        """Seçili aidatları Excel'e aktar"""
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, Alignment, PatternFill
        from datetime import datetime
        
        # Excel dosyası oluştur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Aidat Raporu"
        
        # Başlık satırı
        basliklar = ['ID', 'Daire', 'Dönem', 'Aidat Tipi', 'Gider Tipi', 'Tutar', 'Ödendi mi?', 'Ödeme Tarihi', 'Ödeyen Kişi', 'Açıklama']
        for col, baslik in enumerate(basliklar, 1):
            cell = ws.cell(row=1, column=col, value=baslik)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1e3c72", end_color="1e3c72", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Veri satırları
        row = 2
        for aidat in queryset:
            ws.cell(row=row, column=1, value=aidat.id)
            ws.cell(row=row, column=2, value=str(aidat.daire))
            ws.cell(row=row, column=3, value=f"{aidat.ay}/{aidat.yil}")
            ws.cell(row=row, column=4, value=aidat.get_aidat_tipi_display())
            ws.cell(row=row, column=5, value=aidat.gider.get_tip_display() if aidat.gider else "-")
            ws.cell(row=row, column=6, value=float(aidat.tutar))
            ws.cell(row=row, column=7, value="Ödendi" if aidat.odeme_yapildi_mi else "Ödenmedi")
            ws.cell(row=row, column=8, value=aidat.odeme_tarihi.strftime("%d/%m/%Y") if aidat.odeme_tarihi else "-")
            ws.cell(row=row, column=9, value=aidat.kim_odedi.ad_soyad if aidat.kim_odedi else "-")
            ws.cell(row=row, column=10, value=aidat.aciklama)
            row += 1
        
        # Sütun genişliklerini ayarla
        for col in range(1, 11):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        
        # Response oluştur
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=aidat_raporu_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response
    aidat_raporu_excel.short_description = "Seçili aidatları Excel'e aktar"
    
    def changelist_view(self, request, extra_context=None):
        from django.db.models import Sum
        queryset = self.get_queryset(request)
        
        # Filtre uygulandı mı kontrol et
        try:
            from django.contrib.admin.views.main import ChangeList
            cl = ChangeList(request, self.model, self.list_display, self.list_display_links,
                            self.list_filter, self.date_hierarchy, self.search_fields,
                            self.list_select_related, self.list_per_page, self.list_max_show_all,
                            self.list_editable, self, self.sortable_by)
            queryset = cl.get_queryset(request)
        except:
            pass
        
        toplam_aidat = queryset.aggregate(Sum('tutar'))['tutar__sum'] or 0
        toplam_odenmis = queryset.filter(odeme_yapildi_mi=True).aggregate(Sum('tutar'))['tutar__sum'] or 0
        toplam_odenmemis = queryset.filter(odeme_yapildi_mi=False).aggregate(Sum('tutar'))['tutar__sum'] or 0
        
        # Gider tipine göre istatistik
        gider_istatistik = []
        if 'gider__tip' in request.GET:
            gider_tipi = request.GET.get('gider__tip')
            if gider_tipi:
                gider_aidat = queryset.filter(gider__tip=gider_tipi)
                gider_istatistik = {
                    'tip': gider_tipi,
                    'toplam': gider_aidat.aggregate(Sum('tutar'))['tutar__sum'] or 0,
                    'odenmis': gider_aidat.filter(odeme_yapildi_mi=True).aggregate(Sum('tutar'))['tutar__sum'] or 0,
                    'sayi': gider_aidat.count(),
                    'odenmis_sayi': gider_aidat.filter(odeme_yapildi_mi=True).count(),
                }
        
        extra_context = extra_context or {}
        extra_context['toplam_aidat'] = toplam_aidat
        extra_context['toplam_odenmis'] = toplam_odenmis
        extra_context['toplam_odenmemis'] = toplam_odenmemis
        extra_context['odenme_orani'] = (toplam_odenmis / toplam_aidat * 100) if toplam_aidat > 0 else 0
        extra_context['gider_istatistik'] = gider_istatistik
        
        return super().changelist_view(request, extra_context=extra_context)

class RaporlarAdmin(admin.AdminSite):
    def index(self, request, extra_context=None):
        """Ana sayfaya rapor menüsünü ekle"""
        extra_context = extra_context or {}
        extra_context['raporlar'] = [
            {'name': '📊 Rapor Ana Sayfa', 'url': 'raporlar/'},
            {'name': '💰 Gelir - Gider Raporu', 'url': 'raporlar/gelir-gider/'},
            {'name': '🏦 Banka Hareketleri Raporu', 'url': 'raporlar/banka-hareketleri/'},
            {'name': '📈 Aidat Durumu Raporu', 'url': 'raporlar/aidat-durumu/'},
            {'name': '📋 Genel Durum Raporu', 'url': 'raporlar/genel-durum/'},
        ]
        return super().index(request, extra_context)
    
    def api_daireler(self, request):
        from django.http import JsonResponse
        from bina.models import Daire
        
        blok_id = request.GET.get('blok')
        if blok_id:
            daireler = Daire.objects.filter(blok_id=blok_id).order_by('daire_no')
            data = [{'id': d.id, 'ad': f"{d.blok.get_blok_adi_display()}-{d.daire_no}"} for d in daireler]
            return JsonResponse(data, safe=False)
        return JsonResponse([], safe=False)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('raporlar/', self.admin_view(self.rapor_ana_sayfa), name='rapor_ana_sayfa'),
            path('raporlar/gelir-gider/', self.admin_view(self.gelir_gider_raporu), name='gelir_gider_raporu'),
            path('raporlar/banka-hareketleri/', self.admin_view(self.banka_hareketleri_raporu), name='banka_hareketleri_raporu'),
            path('raporlar/aidat-durumu/', self.admin_view(self.aidat_durumu_raporu), name='aidat_durumu_raporu'),
            path('raporlar/genel-durum/', self.admin_view(self.genel_durum_raporu), name='genel_durum_raporu'),
            path('raporlar/maas-bordrosu/', self.admin_view(self.maas_bordrosu_raporu), name='maas_bordrosu_raporu'),  # <-- YENİ
            path('bina/fatura/upload/', self.admin_view(self.fatura_upload), name='fatura_upload'),
            path('bina/banka-hareket/import-excel/', self.admin_view(self.import_banka_hareket_excel), name='import_banka_hareket_excel'),
            path('yedekle/', self.admin_view(self.veritabani_yedekle), name='veritabani_yedekle'),
            path('yedekle/yukle/', self.admin_view(self.yedekten_yukle), name='yedekten_yukle'),
            path('api/daireler/', self.admin_view(self.api_daireler), name='api_daireler'),
        ]
        return custom_urls + urls
    
    def veritabani_yedekle(self, request):
        """Veritabanını ve medya dosyalarını yedekler"""
        # Yedekleme klasörü
        backup_dir = os.path.join(settings.BASE_DIR, 'yedekler')
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        # Zaman damgalı dosya adı
        zaman = datetime.now().strftime('%Y%m%d_%H%M%S')
        zip_adı = f"site_yedek_{zaman}.zip"
        zip_yolu = os.path.join(backup_dir, zip_adı)
        
        # Zip dosyası oluştur
        with zipfile.ZipFile(zip_yolu, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # 1. SQLite veritabanını ekle
            db_path = settings.DATABASES['default']['NAME']
            if os.path.exists(db_path):
                zipf.write(db_path, 'veritabani/db.sqlite3')
            
            # 2. Media dosyalarını ekle (faturalar, dekontlar vb.)
            media_root = settings.MEDIA_ROOT
            if os.path.exists(media_root):
                for root, dirs, files in os.walk(media_root):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, settings.BASE_DIR)
                        zipf.write(file_path, arcname)
            
            # 3. Static dosyalarını ekle (opsiyonel, büyük olabilir)
            # static_root = settings.STATIC_ROOT
            # if os.path.exists(static_root):
            #     for root, dirs, files in os.walk(static_root):
            #         for file in files:
            #             file_path = os.path.join(root, file)
            #             arcname = os.path.relpath(file_path, settings.BASE_DIR)
            #             zipf.write(file_path, arcname)
        
        # Dosyayı indir
        with open(zip_yolu, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_adı}"'
        
        # İndirildikten sonra zip dosyasını sil (isteğe bağlı)
        os.remove(zip_yolu)
        
        return response
    
    
    def yedekten_yukle(self, request):
        """Yedek dosyasından veritabanını geri yükler"""
        if request.method == 'POST':
            dosya = request.FILES.get('yedek_dosya')
            
            if not dosya:
                messages.error(request, 'Lütfen bir yedek dosyası seçin!')
                return render(request, 'admin/yedek_yukle.html', {'title': 'Yedekten Yükle'})
            
            if not dosya.name.endswith('.zip'):
                messages.error(request, 'Lütfen .zip formatında bir dosya seçin!')
                return render(request, 'admin/yedek_yukle.html', {'title': 'Yedekten Yükle'})
            
            try:
                # Geçici klasör oluştur
                with tempfile.TemporaryDirectory() as temp_dir:
                    zip_yolu = os.path.join(temp_dir, 'yedek.zip')
                    
                    # Dosyayı kaydet
                    with open(zip_yolu, 'wb') as f:
                        for chunk in dosya.chunks():
                            f.write(chunk)
                    
                    # Zip dosyasını aç
                    with zipfile.ZipFile(zip_yolu, 'r') as zipf:
                        zipf.extractall(temp_dir)
                    
                    # Veritabanı dosyasını bul
                    db_path = settings.DATABASES['default']['NAME']
                    
                    # Geçici yedek oluştur (güvenlik için)
                    yedek_adi = f"{db_path}.backup"
                    if os.path.exists(db_path):
                        shutil.copy2(db_path, yedek_adi)
                    
                    # Yeni veritabanını kopyala
                    for root, dirs, files in os.walk(temp_dir):
                        for file in files:
                            if file == 'db.sqlite3':
                                kaynak = os.path.join(root, file)
                                shutil.copy2(kaynak, db_path)
                                break
                
                messages.success(request, 'Veritabanı başarıyla geri yüklendi! Sayfa yenileniyor...')
                
                # Yenileme için JavaScript
                return render(request, 'admin/yedek_yukle.html', {
                    'title': 'Yedekten Yükle',
                    'success': True
                })
                
            except Exception as e:
                messages.error(request, f'Yükleme sırasında hata oluştu: {str(e)}')
                return render(request, 'admin/yedek_yukle.html', {'title': 'Yedekten Yükle'})
        
        return render(request, 'admin/yedek_yukle.html', {'title': 'Yedekten Yükle'})

    def import_banka_hareket_excel(self, request):
        """Excel'den banka hareketlerini içe aktar (pandas olmadan)"""
        if request.method == 'POST':
            import openpyxl
            from io import BytesIO
            from datetime import datetime as dt
            
            dosya = request.FILES.get('excel_dosya')
            if not dosya:
                return JsonResponse({'status': 'error', 'message': 'Dosya seçilmedi!'})
            
            # Dosya adını kontrol et
            if not dosya.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({'status': 'error', 'message': 'Lütfen .xlsx veya .xls formatında dosya yükleyin!'})
            
            try:
                # Excel dosyasını openpyxl ile oku
                workbook = openpyxl.load_workbook(BytesIO(dosya.read()))
                sheet = workbook.active
                
                banka_id = request.POST.get('banka')
                if not banka_id:
                    return JsonResponse({'status': 'error', 'message': 'Banka hesabı seçilmedi!'})
                
                try:
                    from .models import Banka
                    banka = Banka.objects.get(id=banka_id)
                except Exception as e:
                    return JsonResponse({'status': 'error', 'message': f'Banka bulunamadı: {str(e)}'})
                
                eklenen = 0
                hatalar = []
                
                # Sütun başlıklarını bul (ilk satır)
                headers = {}
                for col_idx, cell in enumerate(sheet[1], 1):
                    if cell.value:
                        headers[str(cell.value).lower().strip()] = col_idx
                
                # Gerekli sütunları kontrol et
                # Alternatif sütun isimleri
                tarih_sutun = None
                tutar_sutun = None
                aciklama_sutun = None
                
                for key, col in headers.items():
                    if key in ['tarih', 'date', 'tarihi']:
                        tarih_sutun = col
                    if key in ['tutar', 'miktar', 'amount', 'tutarı']:
                        tutar_sutun = col
                    if key in ['aciklama', 'açıklama', 'description', 'explanation']:
                        aciklama_sutun = col
                
                if not tarih_sutun:
                    return JsonResponse({'status': 'error', 'message': 'Excelde "tarih" sütunu bulunamadı!'})
                if not tutar_sutun:
                    return JsonResponse({'status': 'error', 'message': 'Excelde "tutar" sütunu bulunamadı!'})
                
                # Verileri oku (2. satırdan itibaren)
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):
                    try:
                        # Tarih
                        tarih_cell = row[tarih_sutun - 1]
                        if not tarih_cell.value:
                            continue
                        
                        # Tarih dönüşümü
                        if isinstance(tarih_cell.value, str):
                            try:
                                tarih = dt.strptime(tarih_cell.value, '%Y-%m-%d').date()
                            except:
                                try:
                                    tarih = dt.strptime(tarih_cell.value, '%d.%m.%Y').date()
                                except:
                                    try:
                                        tarih = dt.strptime(tarih_cell.value, '%d/%m/%Y').date()
                                    except:
                                        hatalar.append(f"Satır {row_idx}: Tarih formatı hatalı - {tarih_cell.value}")
                                        continue
                        else:
                            tarih = tarih_cell.value.date() if hasattr(tarih_cell.value, 'date') else tarih_cell.value
                        
                        # Tutar
                        tutar_cell = row[tutar_sutun - 1]
                        if not tutar_cell.value:
                            continue
                        
                        try:
                            tutar = float(tutar_cell.value)
                        except:
                            hatalar.append(f"Satır {row_idx}: Tutar formatı hatalı - {tutar_cell.value}")
                            continue
                        
                        # Hareket tipi (pozitif=gelir, negatif=gider)
                        hareket_tipi = 'gelir' if tutar > 0 else 'gider'
                        tutar_abs = abs(tutar)
                        
                        # Açıklama
                        aciklama = ''
                        if aciklama_sutun and len(row) > aciklama_sutun - 1 and row[aciklama_sutun - 1].value:
                            aciklama = str(row[aciklama_sutun - 1].value)
                        
                        # Benzersiz kontrol - aynı kayıt varsa atla
                        existing = BankaHareket.objects.filter(
                            banka=banka,
                            tarih=tarih,
                            tutar=tutar_abs,
                            aciklama=aciklama
                        ).first()
                        
                        if not existing:
                            BankaHareket.objects.create(
                                banka=banka,
                                hareket_tipi=hareket_tipi,
                                tutar=tutar_abs,
                                tarih=tarih,
                                aciklama=aciklama
                            )
                            eklenen += 1
                    
                    except Exception as e:
                        hatalar.append(f"Satır {row_idx}: {str(e)}")
                
                if eklenen == 0 and not hatalar:
                    return JsonResponse({'status': 'error', 'message': 'Excel dosyasında işlenecek veri bulunamadı! İlk satır başlık olmalı.'})
                
                mesaj = f"✅ {eklenen} hareket başarıyla eklendi."
                if hatalar:
                    mesaj += f" ⚠️ {len(hatalar)} satır atlandı."
                
                return JsonResponse({'status': 'success', 'message': mesaj})
                
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'Dosya okuma hatası: {str(e)}'})
        
        return JsonResponse({'status': 'error', 'message': 'Geçersiz istek!'})

    
    def rapor_ana_sayfa(self, request):
        context = {
            'title': 'Raporlar',
            'site_header': self.site_header,
            'site_title': self.site_title,
        }
        return render(request, 'admin/rapor_ana_sayfa.html', context)
    
    def gelir_gider_raporu(self, request):
        yil = request.GET.get('yil', datetime.now().year)
        ay = request.GET.get('ay', None)
        hizli_tarih = request.GET.get('hizli_tarih', None)
        baslangic_tarih = request.GET.get('baslangic_tarih', '')
        bitis_tarih = request.GET.get('bitis_tarih', '')
        
        try:
            yil = int(yil)
        except:
            yil = datetime.now().year
        
        # Hızlı tarih seçimleri
        bugun = datetime.now().date()
        if hizli_tarih == 'son7gun':
            baslangic = bugun - timedelta(days=7)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son15gun':
            baslangic = bugun - timedelta(days=15)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son30gun':
            baslangic = bugun - timedelta(days=30)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son1ay':
            baslangic = bugun - timedelta(days=30)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son3ay':
            baslangic = bugun - timedelta(days=90)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son6ay':
            baslangic = bugun - timedelta(days=180)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son1yil':
            baslangic = bugun - timedelta(days=365)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif baslangic_tarih and bitis_tarih:
            try:
                baslangic = datetime.strptime(baslangic_tarih, '%Y-%m-%d').date()
                bitis = datetime.strptime(bitis_tarih, '%Y-%m-%d').date()
            except:
                baslangic = None
                bitis = None
        elif ay:
            baslangic = None
            bitis = None
        else:
            baslangic = None
            bitis = None
        
        # Tarih aralığına göre filtreleme
        if baslangic and bitis:
            giderler = Gider.objects.filter(tarih__range=[baslangic, bitis])
            depozito_gelirleri = BankaHareket.objects.filter(
                hareket_tipi='gelir',
                depozito__isnull=False,
                tarih__range=[baslangic, bitis]
            )
            aidatlar = Aidat.objects.filter(odeme_tarihi__range=[baslangic, bitis], odeme_yapildi_mi=True)
            donem = f"{baslangic.strftime('%d/%m/%Y')} - {bitis.strftime('%d/%m/%Y')}"
        elif ay:
            giderler = Gider.objects.filter(tarih__year=yil, tarih__month=ay)
            depozito_gelirleri = BankaHareket.objects.filter(
                hareket_tipi='gelir',
                depozito__isnull=False,
                tarih__year=yil,
                tarih__month=ay
            )
            aidatlar = Aidat.objects.filter(yil=yil, ay=ay, odeme_yapildi_mi=True)
            donem = f"{ay}/{yil}"
        else:
            giderler = Gider.objects.filter(tarih__year=yil)
            depozito_gelirleri = BankaHareket.objects.filter(
                hareket_tipi='gelir',
                depozito__isnull=False,
                tarih__year=yil
            )
            aidatlar = Aidat.objects.filter(yil=yil, odeme_yapildi_mi=True)
            donem = f"{yil} Yılı"
        
        toplam_aidat_geliri = aidatlar.aggregate(Sum('tutar'))['tutar__sum'] or 0
        toplam_depozito_geliri = depozito_gelirleri.aggregate(Sum('tutar'))['tutar__sum'] or 0
        toplam_gider = giderler.aggregate(Sum('tutar'))['tutar__sum'] or 0
        gider_kategorileri = giderler.values('tip').annotate(toplam=Sum('tutar'))
        
        # Aylar için Türkçe isimler
        aylar = {
            '1': 'Ocak', '2': 'Şubat', '3': 'Mart', '4': 'Nisan',
            '5': 'Mayıs', '6': 'Haziran', '7': 'Temmuz', '8': 'Ağustos',
            '9': 'Eylül', '10': 'Ekim', '11': 'Kasım', '12': 'Aralık'
        }
        
        context = {
            'title': 'Gelir - Gider Raporu',
            'site_header': self.site_header,
            'yil': yil,
            'ay': ay,
            'ay_adi': aylar.get(ay, ''),
            'hizli_tarih': hizli_tarih,
            'baslangic_tarih': baslangic_tarih,
            'bitis_tarih': bitis_tarih,
            'donem': donem,
            'toplam_aidat_geliri': toplam_aidat_geliri,
            'toplam_depozito_geliri': toplam_depozito_geliri,
            'toplam_gelir': float(toplam_aidat_geliri) + float(toplam_depozito_geliri),
            'toplam_gider': toplam_gider,
            'net_durum': float(toplam_aidat_geliri) + float(toplam_depozito_geliri) - float(toplam_gider),
            'gider_kategorileri': gider_kategorileri,
        }
        return render(request, 'admin/gelir_gider_raporu.html', context)
    
    def banka_hareketleri_raporu(self, request):
        yil = request.GET.get('yil', datetime.now().year)
        ay = request.GET.get('ay', None)
        banka_id = request.GET.get('banka', None)
        hizli_tarih = request.GET.get('hizli_tarih', None)
        baslangic_tarih = request.GET.get('baslangic_tarih', '')
        bitis_tarih = request.GET.get('bitis_tarih', '')
        
        try:
            yil = int(yil)
        except:
            yil = datetime.now().year
        
        # Hızlı tarih seçimleri
        bugun = datetime.now().date()
        if hizli_tarih == 'son7gun':
            baslangic = bugun - timedelta(days=7)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih in ['son15gun', 'son30gun', 'son1ay']:
            gun = 15 if hizli_tarih == 'son15gun' else 30
            baslangic = bugun - timedelta(days=gun)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son3ay':
            baslangic = bugun - timedelta(days=90)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son6ay':
            baslangic = bugun - timedelta(days=180)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son1yil':
            baslangic = bugun - timedelta(days=365)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif baslangic_tarih and bitis_tarih:
            try:
                baslangic = datetime.strptime(baslangic_tarih, '%Y-%m-%d').date()
                bitis = datetime.strptime(bitis_tarih, '%Y-%m-%d').date()
            except:
                baslangic = None
                bitis = None
        else:
            baslangic = None
            bitis = None
        
        # Tarih aralığına göre filtreleme
        if baslangic and bitis:
            hareketler = BankaHareket.objects.filter(tarih__range=[baslangic, bitis])
            if banka_id:
                hareketler = hareketler.filter(banka_id=banka_id)
            donem = f"{baslangic.strftime('%d/%m/%Y')} - {bitis.strftime('%d/%m/%Y')}"
        elif ay:
            hareketler = BankaHareket.objects.filter(tarih__year=yil, tarih__month=ay)
            if banka_id:
                hareketler = hareketler.filter(banka_id=banka_id)
            donem = f"{ay}/{yil}"
        else:
            hareketler = BankaHareket.objects.filter(tarih__year=yil)
            if banka_id:
                hareketler = hareketler.filter(banka_id=banka_id)
            donem = f"{yil} Yılı"
        
        toplam_gelir = hareketler.filter(hareket_tipi='gelir').aggregate(Sum('tutar'))['tutar__sum'] or 0
        toplam_gider = hareketler.filter(hareket_tipi='gider').aggregate(Sum('tutar'))['tutar__sum'] or 0
        bankalar = Banka.objects.all()
        
        # Aylar için Türkçe isimler
        aylar = {1: 'Ocak', 2: 'Şubat', 3: 'Mart', 4: 'Nisan', 5: 'Mayıs', 6: 'Haziran',
                7: 'Temmuz', 8: 'Ağustos', 9: 'Eylül', 10: 'Ekim', 11: 'Kasım', 12: 'Aralık'}
        
        context = {
            'title': 'Banka Hareketleri Raporu',
            'site_header': self.site_header,
            'yil': yil,
            'ay': ay,
            'ay_adi': aylar.get(ay, ''),
            'banka_id': banka_id,
            'hizli_tarih': hizli_tarih,
            'baslangic_tarih': baslangic_tarih,
            'bitis_tarih': bitis_tarih,
            'donem': donem,
            'hareketler': hareketler.order_by('-tarih'),
            'toplam_gelir': toplam_gelir,
            'toplam_gider': toplam_gider,
            'net_bakiye': toplam_gelir - toplam_gider,
            'bankalar': bankalar,
        }
        return render(request, 'admin/banka_hareketleri_raporu.html', context)
    
    def aidat_durumu_raporu(self, request):
        yil = request.GET.get('yil', datetime.now().year)
        ay = request.GET.get('ay', None)
        blok_id = request.GET.get('blok', None)      # YENİ
        daire_id = request.GET.get('daire', None)    # YENİ
        odeme_durumu = request.GET.get('odeme_durumu', None)  # YENİ
        aidat_tipi = request.GET.get('aidat_tipi', None)      # YENİ
        hizli_tarih = request.GET.get('hizli_tarih', None)
        baslangic_tarih = request.GET.get('baslangic_tarih', '')
        bitis_tarih = request.GET.get('bitis_tarih', '')
        
        try:
            yil = int(yil)
        except:
            yil = datetime.now().year
        
        # Hızlı tarih seçimleri
        bugun = datetime.now().date()
        if hizli_tarih == 'son7gun':
            baslangic = bugun - timedelta(days=7)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih in ['son15gun', 'son30gun', 'son1ay']:
            gun = 15 if hizli_tarih == 'son15gun' else 30
            baslangic = bugun - timedelta(days=gun)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son3ay':
            baslangic = bugun - timedelta(days=90)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son6ay':
            baslangic = bugun - timedelta(days=180)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif hizli_tarih == 'son1yil':
            baslangic = bugun - timedelta(days=365)
            bitis = bugun
            baslangic_tarih = baslangic.strftime('%Y-%m-%d')
            bitis_tarih = bitis.strftime('%Y-%m-%d')
        elif baslangic_tarih and bitis_tarih:
            try:
                baslangic = datetime.strptime(baslangic_tarih, '%Y-%m-%d').date()
                bitis = datetime.strptime(bitis_tarih, '%Y-%m-%d').date()
            except:
                baslangic = None
                bitis = None
        else:
            baslangic = None
            bitis = None
        
        # Tarih aralığına göre filtreleme
        if baslangic and bitis:
            aidatlar = Aidat.objects.filter(odeme_tarihi__range=[baslangic, bitis])
            donem = f"{baslangic.strftime('%d/%m/%Y')} - {bitis.strftime('%d/%m/%Y')}"
        elif ay:
            aidatlar = Aidat.objects.filter(yil=yil, ay=ay)
            donem = f"{ay}/{yil}"
        else:
            aidatlar = Aidat.objects.filter(yil=yil)
            donem = f"{yil} Yılı"
        
        # ========== YENİ EKLENEN FİLTRELER ==========
        if blok_id:
            aidatlar = aidatlar.filter(daire__blok__id=blok_id)
        if daire_id:
            aidatlar = aidatlar.filter(daire__id=daire_id)
        if odeme_durumu == 'odendi':
            aidatlar = aidatlar.filter(odeme_yapildi_mi=True)
        elif odeme_durumu == 'odenmedi':
            aidatlar = aidatlar.filter(odeme_yapildi_mi=False)
        if aidat_tipi:
            aidatlar = aidatlar.filter(aidat_tipi=aidat_tipi)
        # ===========================================
        
        odenen = aidatlar.filter(odeme_yapildi_mi=True).aggregate(Sum('tutar'))['tutar__sum'] or 0
        odenmeyen = aidatlar.filter(odeme_yapildi_mi=False).aggregate(Sum('tutar'))['tutar__sum'] or 0
        toplam = odenen + odenmeyen
        
        # Daire bazlı durum - Filtrelere göre daireleri göster
        if daire_id:
            daireler = Daire.objects.filter(id=daire_id)
        elif blok_id:
            daireler = Daire.objects.filter(blok__id=blok_id)
        else:
            daireler = Daire.objects.all()
        
        daire_durum = []
        for daire in daireler:
            daire_aidat = aidatlar.filter(daire=daire)
            odenen_borc = daire_aidat.filter(odeme_yapildi_mi=True).aggregate(Sum('tutar'))['tutar__sum'] or 0
            toplam_borc = daire_aidat.aggregate(Sum('tutar'))['tutar__sum'] or 0
            kalan = toplam_borc - odenen_borc
            
            if toplam_borc > 0 or daire_aidat.exists():
                # Durum belirleme
                if kalan == 0:
                    durum = "✅ Ödendi"
                elif odenen_borc > 0:
                    durum = "⚠️ Kısmi Ödendi"
                else:
                    durum = "❌ Ödenmedi"
                
                daire_durum.append({
                    'daire': f"{daire.blok.get_blok_adi_display()}-{daire.daire_no}",
                    'daire_tipi': daire.get_daire_tipi_display(),
                    'toplam_borc': toplam_borc,
                    'odenen': odenen_borc,
                    'kalan': kalan,
                    'oran': (odenen_borc / toplam_borc * 100) if toplam_borc > 0 else 0,
                    'aidat_sayisi': daire_aidat.count(),
                    'odenen_sayisi': daire_aidat.filter(odeme_yapildi_mi=True).count(),
                    'durum': durum,  # ← BURASI ÖNEMLİ
                })
        
        aylar = {1: 'Ocak', 2: 'Şubat', 3: 'Mart', 4: 'Nisan', 5: 'Mayıs', 6: 'Haziran',
                7: 'Temmuz', 8: 'Ağustos', 9: 'Eylül', 10: 'Ekim', 11: 'Kasım', 12: 'Aralık'}
        
        # Filtreler için blok listesi
        from bina.models import Blok
        bloklar = Blok.objects.all()
        
        context = {
            'title': 'Aidat Durumu Raporu',
            'site_header': self.site_header,
            'yil': yil,
            'ay': ay,
            'ay_adi': aylar.get(int(ay), '') if ay else '',
            'blok_id': blok_id,
            'daire_id': daire_id,
            'odeme_durumu': odeme_durumu,
            'aidat_tipi': aidat_tipi,
            'hizli_tarih': hizli_tarih,
            'baslangic_tarih': baslangic_tarih,
            'bitis_tarih': bitis_tarih,
            'donem': donem,
            'toplam': toplam,
            'odenen': odenen,
            'odenen_orani': (odenen / toplam * 100) if toplam > 0 else 0,
            'odunmeyen': odenmeyen,
            'daire_durum': daire_durum,
            'bloklar': bloklar,
            'secili_blok': Blok.objects.filter(id=blok_id).first() if blok_id else None,
            'secili_daire': Daire.objects.filter(id=daire_id).first() if daire_id else None,
        }
        return render(request, 'admin/aidat_durumu_raporu.html', context)
    
    def maas_bordrosu_raporu(self, request):
        from .models import MaasBordrosu
        from calendar import month_name
        from django.db.models import Sum
        
        yil = request.GET.get('yil', None)
        ay = request.GET.get('ay', None)
        
        if not yil:
            from datetime import datetime
            yil = datetime.now().year
        
        try:
            yil = int(yil)
        except:
            yil = datetime.now().year
        
        if ay:
            try:
                ay = int(ay)
            except:
                ay = None
        
        # Bordroları filtrele
        bordrolar = MaasBordrosu.objects.all()
        if ay:
            bordrolar = bordrolar.filter(yil=yil, ay=ay)
            donem = f"{month_name[ay]} {yil}"
        else:
            bordrolar = bordrolar.filter(yil=yil)
            donem = f"{yil} Yılı"
        
        # Toplamlar
        toplam_brut = bordrolar.aggregate(Sum('brut_maas'))['brut_maas__sum'] or 0
        toplam_net = bordrolar.aggregate(Sum('net_maas'))['net_maas__sum'] or 0
        toplam_maliyet = bordrolar.aggregate(Sum('isveren_toplam_maliyet'))['isveren_toplam_maliyet__sum'] or 0
        toplam_fazla_mesai = bordrolar.aggregate(Sum('fazla_mesai_tutari'))['fazla_mesai_tutari__sum'] or 0
        toplam_tatil = bordrolar.aggregate(Sum('resmi_tatil_tutari'))['resmi_tatil_tutari__sum'] or 0
        toplam_bayram = bordrolar.aggregate(Sum('bayram_tutari'))['bayram_tutari__sum'] or 0
        toplam_prim = bordrolar.aggregate(Sum('prim'))['prim__sum'] or 0
        toplam_sgk = bordrolar.aggregate(Sum('sgk_isci_payi'))['sgk_isci_payi__sum'] or 0
        toplam_issizlik = bordrolar.aggregate(Sum('issizlik_isci_payi'))['issizlik_isci_payi__sum'] or 0
        toplam_vergi = bordrolar.aggregate(Sum('gelir_vergisi'))['gelir_vergisi__sum'] or 0
        toplam_damga = bordrolar.aggregate(Sum('damga_vergisi'))['damga_vergisi__sum'] or 0
        toplam_sgk_isveren = bordrolar.aggregate(Sum('sgk_isveren_payi'))['sgk_isveren_payi__sum'] or 0
        toplam_issizlik_isveren = bordrolar.aggregate(Sum('issizlik_isveren_payi'))['issizlik_isveren_payi__sum'] or 0
        
        aylar = [(1, 'Ocak'), (2, 'Şubat'), (3, 'Mart'), (4, 'Nisan'), 
                (5, 'Mayıs'), (6, 'Haziran'), (7, 'Temmuz'), (8, 'Ağustos'),
                (9, 'Eylül'), (10, 'Ekim'), (11, 'Kasım'), (12, 'Aralık')]
        
        yillar = list(range(2020, 2031))
        
        context = {
            'title': 'Maaş Bordrosu Raporu',
            'site_header': self.site_header,
            'donem': donem,
            'bordrolar': bordrolar,
            'yil': yil,
            'ay': ay,
            'yillar': yillar,
            'aylar': aylar,
            'toplam_brut': toplam_brut,
            'toplam_net': toplam_net,
            'toplam_maliyet': toplam_maliyet,
            'toplam_fazla_mesai': toplam_fazla_mesai,
            'toplam_tatil': toplam_tatil,
            'toplam_bayram': toplam_bayram,
            'toplam_prim': toplam_prim,
            'toplam_sgk': toplam_sgk,
            'toplam_issizlik': toplam_issizlik,
            'toplam_vergi': toplam_vergi,
            'toplam_damga': toplam_damga,
            'toplam_sgk_isveren': toplam_sgk_isveren,
            'toplam_issizlik_isveren': toplam_issizlik_isveren,
        }
        return render(request, 'admin/maas_bordrosu_raporu.html', context)

    def genel_durum_raporu(self, request):
        toplam_daire = Daire.objects.count()
        toplam_kisi = Kisi.objects.count()
        toplam_blok = Blok.objects.count()
        bankalar = Banka.objects.all()
        toplam_bakiye = sum([float(b.guncel_bakiye) for b in bankalar])
        odenmemis_aidat = Aidat.objects.filter(odeme_yapildi_mi=False).aggregate(Sum('tutar'))['tutar__sum'] or 0
        bu_yil = datetime.now().year
        yillik_gelir = BankaHareket.objects.filter(tarih__year=bu_yil, hareket_tipi='gelir').aggregate(Sum('tutar'))['tutar__sum'] or 0
        yillik_gider = BankaHareket.objects.filter(tarih__year=bu_yil, hareket_tipi='gider').aggregate(Sum('tutar'))['tutar__sum'] or 0
        depozitolar = Depozito.objects.filter(durum='alindi').aggregate(Sum('tutar'))['tutar__sum'] or 0
        
        context = {
            'title': 'Genel Durum Raporu',
            'site_header': self.site_header,
            'toplam_daire': toplam_daire, 'toplam_kisi': toplam_kisi,
            'toplam_blok': toplam_blok, 'toplam_bakiye': toplam_bakiye,
            'odenmemis_aidat': odenmemis_aidat, 'yillik_gelir': yillik_gelir,
            'yillik_gider': yillik_gider, 'yillik_net': yillik_gelir - yillik_gider,
            'depozitolar': depozitolar, 'bankalar': bankalar,
        }
        return render(request, 'admin/genel_durum_raporu.html', context)
    
    def fatura_upload(self, request):
        """Fatura/dekont dosyası yükleme"""
        if request.method == 'POST':
            from datetime import date
            from django.core.files.storage import default_storage
            from django.core.files.base import ContentFile
            import os
            
            tip = request.POST.get('tip')
            aciklama = request.POST.get('aciklama')
            dosya = request.FILES.get('dosya')
            
            if not dosya:
                return JsonResponse({'status': 'error', 'message': 'Dosya seçilmedi!'})
            
            # Dosya uzantısını kontrol et
            allowed_extensions = ['.pdf', '.jpg', '.png', '.xlsx', '.docx']
            ext = os.path.splitext(dosya.name)[1].lower()
            if ext not in allowed_extensions:
                return JsonResponse({'status': 'error', 'message': 'Desteklenmeyen dosya formatı!'})
            
            # Dosyayı kaydet
            file_path = default_storage.save(f'faturalar/{tip}/{dosya.name}', ContentFile(dosya.read()))
            
            # Veritabanına kaydet
            from .models import Fatura
            Fatura.objects.create(
                tip=tip,
                aciklama=aciklama,
                dosya=file_path,
                tarih=date.today()
            )
            
            return JsonResponse({'status': 'success', 'message': 'Dosya başarıyla yüklendi!'})
        
        return JsonResponse({'status': 'error', 'message': 'Geçersiz istek!'})

    def get_app_list(self, request, app_label=None):
        # Mevcut app_list'i al
        try:
            if app_label:
                app_list = super().get_app_list(request, app_label)
            else:
                app_list = super().get_app_list(request)
        except TypeError:
            app_list = super().get_app_list(request)
        
        # Raporlar bölümünü oluştur (sadece ana sayfa için)
        if app_label is None:
            rapor_app = {
                'name': '📊 RAPORLAR',
                'app_label': 'raporlar',
                'app_url': '/admin/raporlar/',
                'models': [
                    {
                        'name': '📊 Rapor Ana Sayfa',
                        'object_name': 'rapor_ana',
                        'admin_url': '/admin/raporlar/',
                        'view_only': True,
                    },
                    {
                        'name': '💰 Gelir - Gider Raporu',
                        'object_name': 'gelir_gider',
                        'admin_url': '/admin/raporlar/gelir-gider/',
                        'view_only': True,
                    },
                    {
                        'name': '🏦 Banka Hareketleri',
                        'object_name': 'banka_hareket',
                        'admin_url': '/admin/raporlar/banka-hareketleri/',
                        'view_only': True,
                    },
                    {
                        'name': '📈 Aidat Durumu',
                        'object_name': 'aidat_durum',
                        'admin_url': '/admin/raporlar/aidat-durumu/',
                        'view_only': True,
                    },
                    {
                        'name': '📋 Genel Durum',
                        'object_name': 'genel_durum',
                        'admin_url': '/admin/raporlar/genel-durum/',
                        'view_only': True,
                    },
                    {
                        'name': '💰 Maaş Bordrosu',
                        'object_name': 'maas_bordrosu',
                        'admin_url': '/admin/raporlar/maas-bordrosu/',
                        'view_only': True,
                    },
                ]
            }
            
            yedekleme_app = {
                'name': '⚙️ SİSTEM',
                'app_label': 'sistem',
                'app_url': '#',
                'models': [
                    {
                        'name': '💾 Veritabanı Yedekle (İndir)',
                        'object_name': 'yedekle',
                        'admin_url': '/admin/yedekle/',
                        'view_only': True,
                    },
                    {
                        'name': '📤 Yedekten Yükle',
                        'object_name': 'yedek_yukle',
                        'admin_url': '/admin/yedekle/yukle/',
                        'view_only': True,
                    },
                ]
            }
            
            # Raporlar bölümünü en başa ekle
            app_list.insert(0, rapor_app)
            
            # Yedekleme bölümünü en sona ekle
            app_list.append(yedekleme_app)
        
        return app_list

# ==================== PERSONEL / MAAŞ ADMIN ====================

from django.contrib import admin
from django.db.models import Sum
from django.http import HttpResponse
from .models import Personel, AsgariUcret, MaasBordrosu

class PersonelAdmin(admin.ModelAdmin):
    list_display = ['ad_soyad', 'unvan', 'brut_maas', 'maas_tipi', 'calisma_sekli', 'aktif']
    list_filter = ['aktif', 'calisma_sekli', 'sgk_tipi', 'maas_tipi']
    search_fields = ['ad_soyad', 'tc_kimlik']
    fieldsets = (
        ('Kişisel Bilgiler', {
            'fields': ('ad_soyad', 'tc_kimlik', 'unvan', 'ise_baslama_tarihi', 'isten_cikis_tarihi')
        }),
        ('Maaş Bilgileri', {
            'fields': ('maas_tipi', 'brut_maas', 'calisma_sekli', 'sgk_tipi', 'tesvik_durumu', 'aktif')
        }),
        ('Tazminat Bilgileri', {
            'fields': ('ise_giris_tarihi', 'onceki_isveren_hizmeti'),
            'classes': ('collapse',)
        }),
        ('Gider Yuvarlama Ayarları', {
        'fields': ('gider_yuvarlama_aktif', 'gider_yuvarlama_tip', 'gider_yuvarlama_kat', 'depozito_ekle'),
        'classes': ('collapse',)
        }),
    )
    
    class Media:
        js = ('admin/js/personel_maas_tipi.js',)
    
    # Doğrudan form içine JavaScript ekle
    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}
        extra_context['maas_tipi_js'] = """
        <script>
        (function() {
            function updateBrutMaasLabel() {
                var maasTipi = document.getElementById('id_maas_tipi');
                var label = document.querySelector('.field-brut_maas label');
                var input = document.getElementById('id_brut_maas');
                if (maasTipi && label) {
                    if (maasTipi.value === 'net') {
                        label.innerHTML = 'Net Maaş (TL):';
                        if (input) input.placeholder = '30000 TL (Net)';
                    } else {
                        label.innerHTML = 'Brüt Maaş (TL):';
                        if (input) input.placeholder = '33030 TL (Brüt)';
                    }
                }
            }
            document.addEventListener('DOMContentLoaded', updateBrutMaasLabel);
            if (document.getElementById('id_maas_tipi')) {
                document.getElementById('id_maas_tipi').addEventListener('change', updateBrutMaasLabel);
            }
        })();
        </script>
        """
        return super().changeform_view(request, object_id, form_url, extra_context)

class AsgariUcretAdmin(admin.ModelAdmin):
    list_display = ['yil', 'brut_ucret', 'isci_sgk_payi', 'isveren_sgk_payi']
    list_editable = ['brut_ucret']

class GiderAdmin(TarihFiltresiMixin, admin.ModelAdmin):
    tarih_alan = 'tarih'
    list_display = ('tip', 'tutar', 'tarih', 'hesap_tipi', 'blok', 'muaf_daire_sayisi', 'aciklama', 'taksitlendir_button')
    list_filter = ('tip', 'tarih', 'hesap_tipi', 'blok')
    search_fields = ('aciklama', 'fatura_no')
    filter_horizontal = ('muaf_daireler',)  # Muaf daireleri güzel bir arayüzle seçmek için
    actions = ['aidatlari_yeniden_olustur', 'aidatlari_temizle', 'blok_bazli_gider_olustur']
    
    fieldsets = (
        ('Temel Bilgiler', {
            'fields': ('tip', 'tutar', 'tarih', 'hesap_tipi', 'blok')
        }),
        ('Fatura Bilgileri', {
            'fields': ('fatura_no', 'fatura_donemi', 'son_odeme_tarihi', 'odeme_tarihi'),
            'classes': ('collapse',)
        }),
        ('İlişkiler', {
            'fields': ('abonelik', 'firma'),
            'classes': ('collapse',)
        }),
        ('Muafiyet Ayarları', {
            'fields': ('muaf_daireler',),
            'description': 'Bu giderden muaf tutulacak daireleri seçin. Seçili dairelere aidat yansıtılmaz.',
            'classes': ('wide',)
        }),
        ('Açıklama', {
            'fields': ('aciklama',),
        }),
    )
    
    def taksitlendir_button(self, obj):
        from django.utils.html import format_html
        from django.urls import reverse
        url = reverse('admin:gider_taksitlendir', args=[obj.id])
        return format_html('<a href="{}" style="background:#28a745; color:white; padding:5px 10px; border-radius:3px; text-decoration:none;">💰 Taksitlendir</a>', url)
    taksitlendir_button.short_description = 'Taksitlendir'
    taksitlendir_button.allow_tags = True

    def muaf_daire_sayisi(self, obj):
        return obj.muaf_daireler.count()
    muaf_daire_sayisi.short_description = 'Muaf Daire Sayısı'
    
    def aidatlari_yeniden_olustur(self, request, queryset):
        """Seçili giderler için aidatları yeniden oluştur (önce temizle)"""
        from bina.models import Aidat, DepozitoHareket, GiderTaksit
        
        print("=" * 50)
        print("AİDAT YENİDEN OLUŞTURMA BAŞLADI")
        print(f"Seçili gider sayısı: {queryset.count()}")
        print("=" * 50)
        
        for gider in queryset:
            #if gider.hesap_tipi != 'sabit_aidat':
                # Önce eski aidatları ve depozito hareketlerini temizle
                eski_aidat = Aidat.objects.filter(gider=gider).count()
                eski_depo = DepozitoHareket.objects.filter(gider=gider).count()
                eski_taksit = GiderTaksit.objects.filter(gider=gider).count()
                
                Aidat.objects.filter(gider=gider).delete()
                DepozitoHareket.objects.filter(gider=gider).delete()
                GiderTaksit.objects.filter(gider=gider).delete()
                
                print(f"🗑️ {gider}: {eski_aidat} aidat, {eski_depo} depozito, {eski_taksit} taksit temizlendi")
                
                # Yeniden oluştur
                gider.aidatlari_olustur()
                
                print(f"✅ {gider}: Yeniden oluşturuldu")
            #else:
             #   print(f"⚠️ {gider} sabit aidat gideri, atlanıyor...")
        
        self.message_user(request, f"✅ {queryset.count()} gider için aidatlar yeniden oluşturuldu.")

    aidatlari_yeniden_olustur.short_description = "Seçili giderler için aidatları yeniden oluştur (önce temizle)"
    
    def aidatlari_temizle(self, request, queryset):
        """Seçili giderlerin aidatlarını ve depozito hareketlerini temizle"""
        from bina.models import Aidat, DepozitoHareket, GiderTaksit
        
        print("=" * 50)
        print("AİDAT TEMİZLEME BAŞLADI")
        print(f"Seçili gider sayısı: {queryset.count()}")
        print("=" * 50)
        
        temizlenen_aidat = 0
        temizlenen_depozito = 0
        temizlenen_taksit = 0
        
        for gider in queryset:
            #if gider.hesap_tipi != 'sabit_aidat':
                # Aidatları sil
                aidat_count = Aidat.objects.filter(gider=gider).count()
                aidat_silinen = Aidat.objects.filter(gider=gider).delete()
                temizlenen_aidat += aidat_silinen[0]
                
                # Depozito hareketlerini sil
                depo_count = DepozitoHareket.objects.filter(gider=gider).count()
                depo_silinen = DepozitoHareket.objects.filter(gider=gider).delete()
                temizlenen_depozito += depo_silinen[0]
                
                # Taksitleri sil
                taksit_count = GiderTaksit.objects.filter(gider=gider).count()
                taksit_silinen = GiderTaksit.objects.filter(gider=gider).delete()
                temizlenen_taksit += taksit_silinen[0]
                
                print(f"✅ {gider}: {aidat_count} aidat, {depo_count} depozito hareketi, {taksit_count} taksit silindi")
            #else:
             #   print(f"⚠️ {gider} sabit aidat gideri, atlanıyor...")
        
        self.message_user(
            request, 
            f"✅ Temizleme tamamlandı!\n"
            f"   Silinen aidat: {temizlenen_aidat}\n"
            f"   Silinen depozito hareketi: {temizlenen_depozito}\n"
            f"   Silinen taksit: {temizlenen_taksit}"
        )
        
        print(f"\nTOPLAM: {temizlenen_aidat} aidat, {temizlenen_depozito} depozito, {temizlenen_taksit} taksit silindi")
        print("=" * 50)

    aidatlari_temizle.short_description = "Seçili giderlerin aidatlarını ve depozito hareketlerini temizle"

    def blok_bazli_gider_olustur(self, request, queryset):
        """Seçili gideri blok bazlı olarak böl"""
        for gider in queryset:
            if not gider.blok and gider.tip == 'asansor':
                # Her blok için ayrı gider oluştur
                bloklar = Blok.objects.all()
                for blok in bloklar:
                    yeni_gider = Gider.objects.create(
                        tip=gider.tip,
                        tutar=round(gider.tutar / bloklar.count(), 2),
                        tarih=gider.tarih,
                        hesap_tipi=gider.hesap_tipi,
                        blok=blok,
                        aciklama=f"{gider.aciklama} - {blok.get_blok_adi_display()} Blok",
                        fatura_no=gider.fatura_no,
                        son_odeme_tarihi=gider.son_odeme_tarihi,
                    )
                    print(f"✅ {blok.get_blok_adi_display()} Blok için gider oluşturuldu: {yeni_gider.tutar} TL")
                
                # Orijinal gideri sil veya pasif yap
                gider.delete()
                print(f"Orijinal gider silindi.")

    def gideri_taksitlendir(self, request, queryset):
        from django.shortcuts import render
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        from bina.models import GiderTaksit, Aidat, DepozitoHareket
        from decimal import Decimal
        from datetime import date
        import math
        
        print("=" * 60)
        print("🚀 TAKSİTLENDİRME METODU ÇAĞRILDI")
        print(f"Request method: {request.method}")
        print(f"Seçili gider sayısı: {queryset.count()}")
        print("=" * 60)
        
        
        if request.method == 'POST':
            print("POST request geldi!")
            print(f"POST verileri: {request.POST.keys()}")
            
            taksit_tipi = request.POST.get('taksit_tipi', 'esit')
            taksit_sayisi = int(request.POST.get('taksit_sayisi', 3))
            baslangic_ay = int(request.POST.get('baslangic_ay', date.today().month))
            baslangic_yil = int(request.POST.get('baslangic_yil', date.today().year))
            aciklama_ek = ""
            
            print(f"Taksit tipi: {taksit_tipi}")
            print(f"Taksit sayısı: {taksit_sayisi}")
            
            for gider in queryset:
                toplam_tutar = float(gider.tutar)
                taksit_tutarlari = []
                
                if taksit_tipi == 'esit':
                    # Eşit taksit
                    esit_tutar = toplam_tutar / taksit_sayisi
                    for i in range(taksit_sayisi):
                        if i == taksit_sayisi - 1:
                            tutar = round(toplam_tutar - sum(taksit_tutarlari), 2)
                        else:
                            tutar = round(esit_tutar, 2)
                        taksit_tutarlari.append(tutar)
                    print(f"Eşit taksitler: {taksit_tutarlari}")
                    
                elif taksit_tipi == 'manuel':
                    # Manuel taksit - önce hidden input'dan al
                    manuel_str = request.POST.get('manuel_taksitler', '')
                    if manuel_str:
                        manuel_list = [float(x.strip()) for x in manuel_str.split(',')]
                        taksit_tutarlari = manuel_list
                    else:
                        # Alternatif: direkt input'lardan al
                        for i in range(1, taksit_sayisi + 1):
                            tutar_str = request.POST.get(f'taksit_{i}', '0')
                            try:
                                tutar = float(tutar_str.replace(',', '.'))
                            except:
                                tutar = 0
                            taksit_tutarlari.append(tutar)
                    print(f"Manuel taksitler: {taksit_tutarlari}")
                    
                elif taksit_tipi == 'yuzde':
                    # Yüzde bazlı taksit
                    yuzde_str = request.POST.get('yuzdeler', '')
                    if yuzde_str:
                        yuzdeler = [float(x.strip()) for x in yuzde_str.split(',')]
                    else:
                        yuzdeler = []
                        for i in range(1, taksit_sayisi + 1):
                            yuzde_str = request.POST.get(f'yuzde_{i}', '0')
                            try:
                                yuzde = float(yuzde_str.replace(',', '.'))
                            except:
                                yuzde = 0
                            yuzdeler.append(yuzde)
                    
                    for i, yuzde in enumerate(yuzdeler):
                        if i == taksit_sayisi - 1:
                            tutar = round(toplam_tutar - sum(taksit_tutarlari), 2)
                        else:
                            tutar = round(toplam_tutar * yuzde / 100, 2)
                        taksit_tutarlari.append(tutar)
                    print(f"Yüzde taksitler: {taksit_tutarlari}")
                
                # Eski kayıtları temizle
                Aidat.objects.filter(gider=gider).delete()
                DepozitoHareket.objects.filter(gider=gider).delete()
                GiderTaksit.objects.filter(gider=gider).delete()
                
                # Taksitleri oluştur
                for i, taksit_tutar in enumerate(taksit_tutarlari):
                    taksit_ay = baslangic_ay + i
                    taksit_yil = baslangic_yil
                    while taksit_ay > 12:
                        taksit_ay -= 12
                        taksit_yil += 1
                    taksit_tarihi = date(taksit_yil, taksit_ay, 1)
                    
                    taksit = GiderTaksit.objects.create(
                        gider=gider,
                        taksit_no=i + 1,
                        tutar=Decimal(str(taksit_tutar)),
                        tarih=taksit_tarihi,
                        aciklama=f"{gider.get_tip_display()} - Taksit {i+1}/{taksit_sayisi} - {taksit_tutar:.2f} TL"
                    )
                    print(f"  ✅ Taksit {i+1}: {taksit_tutar} TL - {taksit_tarihi}")
                    
                    try:
                        self._create_taksit_aidat(gider, taksit, taksit_tarihi, aciklama_ek)
                    except Exception as e:
                        print(f"     ❌ Aidat hatası: {e}")
                
                self.message_user(request, f"✅ {gider} - {taksit_sayisi} taksite ayrıldı")
            
            return HttpResponseRedirect(reverse('admin:bina_gider_changelist'))
        
        # GET request için template göster
        from datetime import datetime
        return render(request, 'admin/gider_taksit.html', {
            'giderler': queryset,
            'now': datetime.now()
        })
    
    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('<int:gider_id>/taksitlendir/', self.admin_site.admin_view(self.taksitlendir_sayfasi), name='gider_taksitlendir'),
        ]
        return custom_urls + urls

    def taksitlendir_sayfasi(self, request, gider_id):
        """Tek bir gider için taksitlendirme sayfası"""
        from bina.models import Gider, GiderTaksit, Aidat, DepozitoHareket
        from decimal import Decimal
        from datetime import date
        from django.shortcuts import render, get_object_or_404
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        
        gider = get_object_or_404(Gider, id=gider_id)
        
        # Debug için yazdır
        print(f"Gider ID: {gider.id}")
        print(f"Gider Tutar: {gider.tutar}")
        print(f"Gider Tip: {gider.get_tip_display()}")
        
        if request.method == 'POST':
            # ... POST işlemleri aynı kalacak ...
            taksit_tipi = request.POST.get('taksit_tipi', 'esit')
            taksit_sayisi = int(request.POST.get('taksit_sayisi', 3))
            baslangic_ay = int(request.POST.get('baslangic_ay', date.today().month))
            baslangic_yil = int(request.POST.get('baslangic_yil', date.today().year))
            
            toplam_tutar = float(gider.tutar)
            taksit_tutarlari = []
            
            if taksit_tipi == 'esit':
                esit_tutar = toplam_tutar / taksit_sayisi
                for i in range(taksit_sayisi):
                    if i == taksit_sayisi - 1:
                        tutar = round(toplam_tutar - sum(taksit_tutarlari), 2)
                    else:
                        tutar = round(esit_tutar, 2)
                    taksit_tutarlari.append(tutar)
                    
            elif taksit_tipi == 'manuel':
                manuel_str = request.POST.get('manuel_taksitler', '')
                if manuel_str:
                    taksit_tutarlari = [float(x) for x in manuel_str.split(',')]
                else:
                    for i in range(1, taksit_sayisi + 1):
                        tutar_str = request.POST.get(f'taksit_{i}', '0')
                        try:
                            tutar = float(tutar_str.replace(',', '.'))
                        except:
                            tutar = 0
                        taksit_tutarlari.append(tutar)
                    
            elif taksit_tipi == 'yuzde':
                yuzde_str = request.POST.get('yuzdeler', '')
                if yuzde_str:
                    yuzdeler = [float(x) for x in yuzde_str.split(',')]
                    for i, yuzde in enumerate(yuzdeler):
                        if i == taksit_sayisi - 1:
                            tutar = round(toplam_tutar - sum(taksit_tutarlari), 2)
                        else:
                            tutar = round(toplam_tutar * yuzde / 100, 2)
                        taksit_tutarlari.append(tutar)
                else:
                    for i in range(1, taksit_sayisi + 1):
                        yuzde_str = request.POST.get(f'yuzde_{i}', '0')
                        try:
                            yuzde = float(yuzde_str.replace(',', '.'))
                        except:
                            yuzde = 0
                        if i == taksit_sayisi:
                            tutar = round(toplam_tutar - sum(taksit_tutarlari), 2)
                        else:
                            tutar = round(toplam_tutar * yuzde / 100, 2)
                        taksit_tutarlari.append(tutar)
            
            # Eski kayıtları temizle
            Aidat.objects.filter(gider=gider).delete()
            DepozitoHareket.objects.filter(gider=gider).delete()
            GiderTaksit.objects.filter(gider=gider).delete()
            
            # Taksitleri oluştur
            for i, taksit_tutar in enumerate(taksit_tutarlari):
                taksit_ay = baslangic_ay + i
                taksit_yil = baslangic_yil
                while taksit_ay > 12:
                    taksit_ay -= 12
                    taksit_yil += 1
                taksit_tarihi = date(taksit_yil, taksit_ay, 1)
                
                taksit = GiderTaksit.objects.create(
                    gider=gider,
                    taksit_no=i + 1,
                    tutar=Decimal(str(taksit_tutar)),
                    tarih=taksit_tarihi,
                    aciklama=f"{gider.get_tip_display()} - Taksit {i+1}/{taksit_sayisi} - {taksit_tutar:.2f} TL"
                )
                
                self._create_taksit_aidat(gider, taksit, taksit_tarihi, "")
            
            self.message_user(request, f"✅ {gider} - {taksit_sayisi} taksite ayrıldı")
            return HttpResponseRedirect(reverse('admin:bina_gider_changelist'))
        
        # GET isteği için form göster
        from datetime import datetime
        return render(request, 'admin/gider_taksit.html', {
            'giderler': [gider],
            'gider_tutar': float(gider.tutar),  # Açıkça tutarı gönder
            'gider_adi': str(gider),
            'now': datetime.now()
        })


    def _create_taksit_aidat(self, gider, taksit, taksit_tarihi, aciklama_ek=""):
        """Tek bir taksit için aidat oluştur"""
        from decimal import Decimal
        import math
        from bina.models import Daire, Aidat, SiteAyarlari
        
        print(f"    → Aidat oluşturuluyor: {gider.get_tip_display()} - Taksit {taksit.taksit_no}")
        
        # Sabit aidat kontrolü
        if gider.hesap_tipi == 'sabit_aidat':
            print(f"      ⚠️ Sabit aidat, atlanıyor")
            return
        
        # Demirbaş gideri kontrolü
        muaf_alan = 'demirbas_giderlerinden_muaf'
        
        # Blok bazlı filtreleme
        if gider.blok:
            daireler = Daire.objects.filter(blok=gider.blok, **{muaf_alan: False})
            toplam_arsa_pay = sum([d.arsa_pay_pay for d in daireler])
            print(f"      Blok: {gider.blok}, Daire sayısı: {daireler.count()}")
        else:
            daireler = Daire.objects.filter(**{muaf_alan: False})
            toplam_arsa_pay = sum([d.arsa_pay_pay for d in daireler])
            print(f"      Tüm bloklar, Daire sayısı: {daireler.count()}")
        
        if toplam_arsa_pay == 0:
            toplam_arsa_pay = 1
        
        site_ayar = SiteAyarlari.objects.first()
        yuvarlama_kat = site_ayar.gider_yuvarlama_kat if site_ayar and site_ayar.gider_yuvarlama_aktif else 0
        yuvarlama_tip = site_ayar.gider_yuvarlama_tip if site_ayar else 'yukari'
        
        olusturulan = 0
        for daire in daireler:
            oran = float(daire.arsa_pay_pay) / toplam_arsa_pay
            tutar = float(taksit.tutar) * oran
            
            if tutar == 0:
                continue
            
            if yuvarlama_kat > 0:
                if yuvarlama_tip == 'yukari':
                    yuvarlanan = math.ceil(tutar / yuvarlama_kat) * yuvarlama_kat
                elif yuvarlama_tip == 'asagi':
                    yuvarlanan = math.floor(tutar / yuvarlama_kat) * yuvarlama_kat
                else:
                    yuvarlanan = round(tutar / yuvarlama_kat) * yuvarlama_kat
            else:
                yuvarlanan = tutar
            
            fark = round(yuvarlanan - tutar, 2)
            
            # Açıklama oluştur
            aciklama = f"{gider.get_tip_display()} - Taksit {taksit.taksit_no}/{gider.taksitler.count()} - {gider.tarih.strftime('%d/%m/%Y')} (Hesap: {tutar:.2f} TL → Ödenecek: {yuvarlanan:.2f} TL)"
            if aciklama_ek:
                aciklama += f" - {aciklama_ek}"
            
            Aidat.objects.create(
                daire=daire,
                ay=taksit.tarih.month,
                yil=taksit.tarih.year,
                aidat_tipi='ekstra',
                gider=gider,
                tutar=Decimal(str(yuvarlanan)),
                yuvarlama_farki=Decimal(str(fark)),
                tahakkuk_tarihi=taksit.tarih,
                aciklama=aciklama
            )
            olusturulan += 1
        
        print(f"      ✅ {olusturulan} aidat oluşturuldu")

    
        
class DepozitoAdmin(admin.ModelAdmin):
    list_display = ('daire', 'kisi', 'tutar', 'alinma_tarihi', 'durum', 'guncel_bakiye', 'hareket_ekle_button')
    list_filter = ('durum', 'alinma_tarihi')
    search_fields = ('daire__blok__blok_adi', 'daire__daire_no', 'kisi__ad_soyad')
    readonly_fields = ('guncel_bakiye',)
    
    def hareket_ekle_button(self, obj):
        from django.utils.html import format_html
        from django.urls import reverse
        url = reverse('admin:depozito_hareket_ekle', args=[obj.id])
        return format_html('<a href="{}" style="background:#28a745; color:white; padding:5px 10px; border-radius:3px; text-decoration:none;">➕ Hareket Ekle</a>', url)
    hareket_ekle_button.short_description = 'İşlem'
    hareket_ekle_button.allow_tags = True

    fieldsets = (
        ('Depozito Bilgileri', {
            'fields': ('daire', 'kisi', 'tutar', 'alinma_tarihi', 'durum', 'banka')
        }),
        ('İade Bilgileri', {
            'fields': ('iade_tarihi', 'iade_tutari'),
            'classes': ('collapse',)
        }),
        ('Açıklama', {
            'fields': ('aciklama',)
        }),
        ('Güncel Bakiye', {
            'fields': ('guncel_bakiye',),
            'classes': ('collapse',)
        }),
    )
    
    def guncel_bakiye(self, obj):
        """Depozito güncel bakiyesini hesapla (ana + hareketler)"""
        from .models import DepozitoHareket
        toplam = float(obj.tutar)
        hareketler = DepozitoHareket.objects.filter(depozito=obj)
        for h in hareketler:
            if h.hareket_tipi == 'ekleme':
                toplam += float(h.tutar)
            elif h.hareket_tipi == 'cikarma':
                toplam -= float(h.tutar)
        return f"{toplam:.2f} TL"
    guncel_bakiye.short_description = "Güncel Bakiye (Hareketler Dahil)"
    
    def get_urls(self):
        from django.urls import path
        from django.shortcuts import redirect
        from django.contrib import messages
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        
        urls = super().get_urls()
        custom_urls = [
            path('<int:depozito_id>/hareket-ekle/', self.admin_site.admin_view(self.depozito_hareket_ekle), name='depozito_hareket_ekle'),
        ]
        return custom_urls + urls
    
    def depozito_hareket_ekle(self, request, depozito_id):
        """Depozito hareketi ekleme sayfası"""
        from .models import Depozito, DepozitoHareket
        from django.shortcuts import render, get_object_or_404
        from django import forms
        from decimal import Decimal
        
        depozito = get_object_or_404(Depozito, id=depozito_id)
        
        class DepozitoHareketForm(forms.Form):
            hareket_tipi = forms.ChoiceField(
                choices=[('ekleme', '➕ Ekleme (Depozitoya ekle)'), ('cikarma', '➖ Çıkarma (Depozitodan düş)')],
                widget=forms.Select(attrs={'class': 'form-control'})
            )
            tutar = forms.DecimalField(
                max_digits=10, 
                decimal_places=2,
                widget=forms.NumberInput(attrs={'class': 'form-control', 'step': '0.01'})
            )
            tarih = forms.DateField(
                widget=forms.DateInput(attrs={'class': 'form-control', 'type': 'date'})
            )
            aciklama = forms.CharField(
                widget=forms.Textarea(attrs={'class': 'form-control', 'rows': 3}),
                required=False
            )
        
        if request.method == 'POST':
            form = DepozitoHareketForm(request.POST)
            if form.is_valid():
                hareket = DepozitoHareket.objects.create(
                    depozito=depozito,
                    hareket_tipi=form.cleaned_data['hareket_tipi'],
                    tutar=form.cleaned_data['tutar'],
                    tarih=form.cleaned_data['tarih'],
                    aciklama=form.cleaned_data['aciklama'] or f"{depozito.daire} - {'Ek depozito' if form.cleaned_data['hareket_tipi'] == 'ekleme' else 'Depozito düşümü'}"
                )
                
                # Yeni bakiyeyi hesapla
                toplam = float(depozito.tutar)
                for h in DepozitoHareket.objects.filter(depozito=depozito):
                    if h.hareket_tipi == 'ekleme':
                        toplam += float(h.tutar)
                    else:
                        toplam -= float(h.tutar)
                
                messages.success(
                    request, 
                    f"✅ Depozito hareketi eklendi!\n"
                    f"   İşlem: {'➕ Ekleme' if hareket.hareket_tipi == 'ekleme' else '➖ Çıkarma'}\n"
                    f"   Tutar: {hareket.tutar} TL\n"
                    f"   Yeni bakiye: {toplam:.2f} TL"
                )
                return HttpResponseRedirect(reverse('admin:bina_depozito_change', args=[depozito.id]))
        else:
            from datetime import date
            form = DepozitoHareketForm(initial={'tarih': date.today()})
        
        context = {
            'title': f'Depozito Hareketi Ekle - {depozito.daire}',
            'depozito': depozito,
            'form': form,
            'opts': self.model._meta,
            'original': depozito,
            'has_change_permission': True,
            'guncel_bakiye': self.guncel_bakiye(depozito),
        }
        return render(request, 'admin/depozito_hareket_ekle.html', context)
    
    actions = ['toplu_depozito_hareketi_ekle']
    
    def toplu_depozito_hareketi_ekle(self, request, queryset):
        """Seçili depozitolara toplu hareket ekle"""
        from django.shortcuts import render
        from django import forms
        from decimal import Decimal
        from datetime import date
        
        class TopluHareketForm(forms.Form):
            hareket_tipi = forms.ChoiceField(
                choices=[('ekleme', '➕ Ekleme (Depozitoya ekle)'), ('cikarma', '➖ Çıkarma (Depozitodan düş)')]
            )
            tutar = forms.DecimalField(max_digits=10, decimal_places=2)
            tarih = forms.DateField(initial=date.today)
            aciklama = forms.CharField(required=False, widget=forms.Textarea)
        
        if 'apply' in request.POST:
            form = TopluHareketForm(request.POST)
            if form.is_valid():
                eklenen = 0
                for depozito in queryset:
                    DepozitoHareket.objects.create(
                        depozito=depozito,
                        hareket_tipi=form.cleaned_data['hareket_tipi'],
                        tutar=form.cleaned_data['tutar'],
                        tarih=form.cleaned_data['tarih'],
                        aciklama=form.cleaned_data['aciklama'] or f"Toplu işlem - {depozito.daire}"
                    )
                    eklenen += 1
                self.message_user(request, f"✅ {eklenen} depozitoya hareket eklendi.")
                return HttpResponseRedirect(request.get_full_path())
        else:
            form = TopluHareketForm()
        
        context = {
            'title': 'Toplu Depozito Hareketi Ekle',
            'queryset': queryset,
            'form': form,
            'opts': self.model._meta,
        }
        return render(request, 'admin/toplu_depozito_hareketi.html', context)
    toplu_depozito_hareketi_ekle.short_description = "Seçili depozitolara toplu hareket ekle"

class DepozitoHareketAdmin(admin.ModelAdmin):
    list_display = ('depozito', 'hareket_tipi', 'tutar', 'tarih', 'aciklama')
    list_filter = ('hareket_tipi', 'tarih')
    search_fields = ('aciklama',)

    def changelist_view(self, request, extra_context=None):
        queryset = self.get_queryset(request)
        try:
            from django.contrib.admin.views.main import ChangeList
            cl = ChangeList(request, self.model, self.list_display, self.list_display_links,
                            self.list_filter, self.date_hierarchy, self.search_fields,
                            self.list_select_related, self.list_per_page, self.list_max_show_all,
                            self.list_editable, self, self.sortable_by)
            queryset = cl.get_queryset(request)
        except:
            pass
        from django.db.models import Sum
        ekleme = queryset.filter(hareket_tipi='ekleme').aggregate(Sum('tutar'))['tutar__sum'] or 0
        cikarma = queryset.filter(hareket_tipi='cikarma').aggregate(Sum('tutar'))['tutar__sum'] or 0
        iade = queryset.filter(hareket_tipi='iade').aggregate(Sum('tutar'))['tutar__sum'] or 0
        extra_context = extra_context or {}
        extra_context['toplam_ekleme'] = ekleme
        extra_context['toplam_cikarma'] = cikarma
        extra_context['toplam_iade'] = iade
        return super().changelist_view(request, extra_context=extra_context)

class MaasBordrosuAdmin(admin.ModelAdmin):
    list_display = ['personel', 'ay_goster', 'yil', 'brut_maas', 'net_maas', 'isveren_toplam_maliyet']
    list_filter = ['yil', 'ay']
    search_fields = ['personel__ad_soyad']
    
    def ay_goster(self, obj):
        aylar = ['', 'Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran', 
                 'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık']
        return aylar[obj.ay]
    ay_goster.short_description = 'Ay'
        
    # FIELDSETS - DÜZELTİLDİ (brut_maas sadece bir yerde)
    fieldsets = (
        ('Temel Bilgiler', {
            'fields': ('personel', 'ay', 'yil')
        }),
        ('Maaş Bilgisi', {
            'fields': ('girilen_net_maas',),  # SADECE girilen_net_maas - brut_maas YOK!
            'description': 'Net maaşlı personel için NET maaş, Brüt maaşlı personel için BRÜT maaş giriniz.'
        }),
        ('Ek Ödemeler (Girilecek Değerler)', {
            'fields': ('fazla_mesai_saati', 'resmi_tatil_gun', 'bayram_gun', 'prim'),
            'classes': ('wide',)
        }),
        ('Hesaplanan Ek Ödemeler (Otomatik)', {
            'fields': ('fazla_mesai_tutari', 'resmi_tatil_tutari', 'bayram_tutari'),
            'classes': ('collapse',)
        }),
        ('Hesaplanan Değerler (Otomatik)', {
            'fields': ('brut_maas', 'sgk_isci_payi', 'issizlik_isci_payi', 'gelir_vergisi', 'damga_vergisi', 
                    'toplam_kesinti', 'net_maas', 'sgk_isveren_payi', 'issizlik_isveren_payi', 
                    'isveren_toplam_maliyet'),
            'classes': ('collapse',)
        }),
    )
    
    readonly_fields = ('fazla_mesai_tutari', 'resmi_tatil_tutari', 'bayram_tutari',
                       'brut_maas', 'sgk_isci_payi', 'issizlik_isci_payi', 'gelir_vergisi', 'damga_vergisi', 
                       'toplam_kesinti', 'net_maas', 'sgk_isveren_payi', 'issizlik_isveren_payi', 
                       'isveren_toplam_maliyet')
    
    def save_model(self, request, obj, form, change):
        from decimal import Decimal, ROUND_HALF_UP
        from .models import AsgariUcret
        
        yil = obj.yil if obj.yil else 2026
        asgari = AsgariUcret.objects.get(yil=yil)
        asgari_brut = Decimal('33030.00')
        asgari_net = Decimal('28075.50')
        
        # 1. EK ÖDEMELERİ HESAPLA
        saatlik_ucret = asgari_brut / Decimal(225)
        gunluk_ucret = asgari_brut / Decimal(30)
        
        obj.fazla_mesai_tutari = (saatlik_ucret * Decimal('1.5') * Decimal(str(obj.fazla_mesai_saati or 0))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        obj.resmi_tatil_tutari = (gunluk_ucret * Decimal('2') * Decimal(str(obj.resmi_tatil_gun or 0))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        obj.bayram_tutari = (gunluk_ucret * Decimal('2.5') * Decimal(str(obj.bayram_gun or 0))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        ek_odemeler = obj.fazla_mesai_tutari + obj.resmi_tatil_tutari + obj.bayram_tutari + (obj.prim or 0)
        
        # 2. NET MAAŞTAN BRÜT HESAPLAMA (DOĞRU FORMÜL)
        if obj.personel.maas_tipi == 'net':
            if obj.girilen_net_maas and obj.girilen_net_maas > 0:
                girilen_net = Decimal(str(obj.girilen_net_maas))
            else:
                girilen_net = Decimal(str(obj.personel.brut_maas))
            
            # DOĞRU FORMÜL: Brüt = Asgari Brüt + (Net Fark × 1.42857)
            # 30.000 TL net için: 28075.50 fark 1924.50 × 1.42857 = 2749.29 + 33030 = 35779.29
            if girilen_net <= asgari_net:
                obj.brut_maas = asgari_brut
            else:
                net_fark = girilen_net - asgari_net
                # 1/0.7 = 1.42857 (çünkü vergi diliminde %30 kesinti var)
                obj.brut_maas = (asgari_brut + (net_fark * Decimal('1.3988'))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            if not obj.brut_maas or obj.brut_maas == 0:
                obj.brut_maas = obj.personel.brut_maas
        
        # 3. TOPLAM BRÜT
        toplam_brut = obj.brut_maas + ek_odemeler
        
        # 4. KESİNTİLER
        sgk_orani = Decimal('0.14')
        sgk = (toplam_brut * sgk_orani).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        issizlik = (toplam_brut * Decimal('0.01')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        if toplam_brut <= asgari_brut:
            gelir_vergisi = Decimal('0')
            damga_vergisi = Decimal('0')
        else:
            asgari_ustu = toplam_brut - asgari_brut
            asgari_ustu_sgk = (asgari_ustu * sgk_orani).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            asgari_ustu_issizlik = (asgari_ustu * Decimal('0.01')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            vergi_matrah = asgari_ustu - asgari_ustu_sgk - asgari_ustu_issizlik
            gelir_vergisi = (vergi_matrah * Decimal('0.15')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            damga_vergisi = (asgari_ustu * Decimal('0.00759')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        toplam_kesinti = sgk + issizlik + gelir_vergisi + damga_vergisi
        net_maas = (toplam_brut - toplam_kesinti).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        # 5. İŞVEREN MALİYETİ
        sgk_isveren_orani = Decimal('0.2075')
        if obj.personel.tesvik_durumu:
            sgk_isveren_orani -= Decimal('0.05')
        sgk_isveren = (toplam_brut * sgk_isveren_orani).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        issizlik_isveren = (toplam_brut * Decimal('0.02')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        isveren_toplam = (toplam_brut + sgk_isveren + issizlik_isveren).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        obj.sgk_isci_payi = sgk
        obj.issizlik_isci_payi = issizlik
        obj.gelir_vergisi = gelir_vergisi
        obj.damga_vergisi = damga_vergisi
        obj.toplam_kesinti = toplam_kesinti
        obj.net_maas = net_maas
        obj.sgk_isveren_payi = sgk_isveren
        obj.issizlik_isveren_payi = issizlik_isveren
        obj.isveren_toplam_maliyet = isveren_toplam
        
        super().save_model(request, obj, form, change)

    def changelist_view(self, request, extra_context=None):
        queryset = self.get_queryset(request)
        try:
            from django.contrib.admin.views.main import ChangeList
            cl = ChangeList(request, self.model, self.list_display, self.list_display_links,
                            self.list_filter, self.date_hierarchy, self.search_fields,
                            self.list_select_related, self.list_per_page, self.list_max_show_all,
                            self.list_editable, self, self.sortable_by)
            queryset = cl.get_queryset(request)
        except:
            pass
        from django.db.models import Sum
        toplam_brut = queryset.aggregate(Sum('brut_maas'))['brut_maas__sum'] or 0
        toplam_net = queryset.aggregate(Sum('net_maas'))['net_maas__sum'] or 0
        extra_context = extra_context or {}
        extra_context['toplam_brut_maas'] = toplam_brut
        extra_context['toplam_net_maas'] = toplam_net
        return super().changelist_view(request, extra_context=extra_context)

# Admin sitesi - BURASI ÖNEMLİ
admin_site = RaporlarAdmin(name='myadmin')
admin_site.site_header = "SİTE YÖNETİM PANELİ"
admin_site.site_title = "SİTE YÖNETİM PANELİ"
admin_site.index_title = "YÖNETİM PANELİNE HOŞ GELDİNİZ"

# Modelleri kaydet - SAKIN @admin.register KULLANMA!
# Modelleri kaydet - ÖNCE AsgariUcret ve MaasBordrosu
admin_site.register(AsgariUcret, AsgariUcretAdmin)
admin_site.register(MaasBordrosu, MaasBordrosuAdmin)
admin_site.register(Personel)

from django.urls import path
from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.contrib import messages
from django.utils.safestring import mark_safe
import random
import string

class DaireKullaniciAdmin(admin.ModelAdmin):
    list_display = ('get_kisi_adi', 'daire', 'telefon', 'email', 'sifre_durumu', 'sifre_belirle', 'aktif')
    list_filter = ('aktif', 'daire__blok', 'sifre_sifirlandi_mi')
    search_fields = ('kisi__ad_soyad', 'daire__daire_no', 'telefon')
    
    def get_kisi_adi(self, obj):
        if obj.kisi:
            return obj.kisi.ad_soyad
        return "-"
    get_kisi_adi.short_description = "Daire Sakini"
    
    def sifre_durumu(self, obj):
        if obj.ilk_giris:
            return mark_safe('<span style="color: #ff9800;">⚠️ İlk Giriş Bekleniyor</span>')
        elif obj.sifre_sifirlandi_mi:
            return mark_safe('<span style="color: #dc3545;">🔄 Şifre Sıfırlandı</span>')
        else:
            return mark_safe('<span style="color: #28a745;">✅ Aktif</span>')
    sifre_durumu.short_description = "Şifre Durumu"

    def sifre_belirle(self, obj):
        if obj.kullanici:
            return mark_safe(f'''
                <button type="button" onclick="sifreBelirle({obj.id})" 
                        style="background:#ff9800; color:white; border:none; 
                        padding:5px 10px; border-radius:4px; cursor:pointer;">
                    🔑 Şifre Belirle
                </button>
            ''')
        return "-"
    sifre_belirle.short_description = "Şifre İşlemleri"
    
    def sifre_goster(self, obj):
        """Şifre sıfırlama linki gönderme butonu"""
        if obj.kullanici and obj.email:
            return mark_safe(f'''
                <button type="button" onclick="sifreSifirlamaLinkiGonder({obj.id})" 
                        style="background:#28a745; color:white; border:none; 
                        padding:5px 10px; border-radius:4px; cursor:pointer;">
                    📧 Link Gönder
                </button>
            ''')
        return "-"
    sifre_goster.short_description = "Şifre İşlemleri"
    
    actions = ['toplu_sifre_sifirla', 'toplu_sifre_mail_gonder']
    
    def toplu_sifre_sifirla(self, request, queryset):
        sifreler = []
        for item in queryset:
            if item.kullanici:
                yeni_sifre = item.sifre_sifirla(admin_request=True)
                sifreler.append(f"{item.kisi.ad_soyad}: {yeni_sifre}")
        
        # Sonuçları göster
        mesaj = "Şifreler sıfırlandı:\n" + "\n".join(sifreler)
        self.message_user(request, mesaj, level='SUCCESS')
    toplu_sifre_sifirla.short_description = "Seçili kullanıcıların şifresini sıfırla"
    
    def toplu_sifre_mail_gonder(self, request, queryset):
        """Seçili kullanıcılara giriş bilgilerini e-posta ile gönder"""
        gonderilen = 0
        for item in queryset:
            if item.kullanici and item.email:
                try:
                    from django.core.mail import send_mail
                    send_mail(
                        'Site Yönetim Portalı - Giriş Bilgileriniz',
                        f'''
Sayın {item.kisi.ad_soyad},

Portal giriş bilgileriniz:

🌐 Giriş Adresi: http://127.0.0.1:8000/portal/login/
👤 Kullanıcı Adı: {item.kullanici.username}
🔑 Şifre: (Mevcut şifrenizi kullanın veya "Şifremi Unuttum" butonuna tıklayın)

Not: Şifrenizi bilmiyorsanız yöneticiden yeni şifre talep edebilirsiniz.

İyi günler dileriz.
                        ''',
                        settings.DEFAULT_FROM_EMAIL,
                        [item.email],
                        fail_silently=False
                    )
                    gonderilen += 1
                except Exception as e:
                    self.message_user(request, f"{item.kisi.ad_soyad} için e-posta gönderilemedi: {str(e)}", level='ERROR')
        
        self.message_user(request, f"{gonderilen} kullanıcıya e-posta gönderildi.", level='SUCCESS')
    toplu_sifre_mail_gonder.short_description = "Seçili kullanıcılara giriş bilgilerini e-posta ile gönder"
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<int:id>/sifre-sifirla/', self.sifre_sifirla_view, name='daire_kullanici_sifre_sifirla'),
            path('<int:id>/sifre-linki-gonder/', self.sifre_linki_gonder, name='daire_kullanici_sifre_linki'),
            path('<int:id>/sifre-belirle/', self.sifre_belirle_view, name='daire_kullanici_sifre_belirle'),
        ]
        return custom_urls + urls

    def sifre_belirle_view(self, request, id):
        obj = get_object_or_404(DaireKullanici, id=id)
        if request.method == 'POST':
            yeni_sifre = request.POST.get('yeni_sifre')
            if yeni_sifre and len(yeni_sifre) >= 6:
                obj.kullanici.set_password(yeni_sifre)
                obj.kullanici.save()
                obj.ilk_giris = False
                obj.sifre_sifirlandi_mi = False
                obj.save()
                messages.success(request, f'{obj.kisi.ad_soyad} için yeni şifre belirlendi!')
                return redirect('admin:bina_dairekullanici_changelist')
            else:
                messages.error(request, 'Şifre en az 6 karakter olmalı!')
        
        return render(request, 'admin/sifre_belirle.html', {'obj': obj})        
    
    def sifre_sifirla_view(self, request, id):
        """Tek kullanıcı için şifre sıfırlama"""
        obj = get_object_or_404(DaireKullanici, id=id)
        if obj.kullanici:
            yeni_sifre = obj.sifre_sifirla(admin_request=True)
            messages.success(request, f'✅ {obj.kisi.ad_soyad} için yeni şifre: {yeni_sifre}')
        else:
            messages.error(request, 'Kullanıcı bulunamadı!')
        return redirect('admin:bina_dairekullanici_changelist')
    
    def sifre_goster_json(self, request, id):
        """Ajax ile şifre gösterme (geçici)"""
        obj = get_object_or_404(DaireKullanici, id=id)
        if obj.kullanici:
            # Yeni geçici şifre oluştur
            yeni_sifre = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
            obj.kullanici.set_password(yeni_sifre)
            obj.kullanici.save()
            obj.sifre_sifirlandi_mi = True
            obj.ilk_giris = True
            obj.save()
            
            return JsonResponse({
                'status': 'success', 
                'sifre': yeni_sifre,
                'kullanici_adi': obj.kullanici.username
            })
        return JsonResponse({'status': 'error', 'sifre': 'Şifre bulunamadı'})
    
    def save_model(self, request, obj, form, change):
        yeni_sifre = None
        if obj.kisi and not obj.kullanici:
            # Kullanıcı oluştur
            kullanici_adi = obj._kullanici_adi_olustur()
            yeni_sifre = obj._olustur_sifre()
            
            user = User.objects.create_user(
                username=kullanici_adi,
                password=yeni_sifre,
                first_name=obj.kisi.ad_soyad.split()[0] if ' ' in obj.kisi.ad_soyad else obj.kisi.ad_soyad,
                last_name=obj.kisi.ad_soyad.split()[-1] if ' ' in obj.kisi.ad_soyad else '',
                email=obj.email
            )
            obj.kullanici = user
            obj.ilk_giris = True
            obj.sifre_sifirlandi_mi = True
            
            if not obj.telefon and obj.kisi.telefon:
                obj.telefon = obj.kisi.telefon
            if not obj.email and obj.kisi.email:
                obj.email = obj.kisi.email
        
        super().save_model(request, obj, form, change)
        
        if yeni_sifre:
            self.message_user(request, 
                             f'✅ Kullanıcı oluşturuldu!<br>'
                             f'👤 Kullanıcı adı: <strong>{obj.kullanici.username}</strong><br>'
                             f'🔑 Şifre: <strong>{yeni_sifre}</strong><br>'
                             f'📧 E-posta: <strong>{obj.email}</strong>',
                             level='SUCCESS')
    
    class Media:
        js = ('admin/js/daire_kullanici.js',)
    
    def sifre_linki_gonder(self, request, id):
        """Kullanıcıya şifre sıfırlama linki gönder (konsola yazdırır)"""
        from django.contrib.auth.tokens import default_token_generator
        from django.utils.http import urlsafe_base64_encode
        from django.utils.encoding import force_bytes
        from django.core.mail import send_mail
        
        obj = get_object_or_404(DaireKullanici, id=id)
        
        if not obj.kullanici or not obj.email:
            messages.error(request, 'Kullanıcı veya e-posta adresi bulunamadı!')
            return redirect('admin:bina_dairekullanici_changelist')
        
        # Şifre sıfırlama linki oluştur
        uid = urlsafe_base64_encode(force_bytes(obj.kullanici.pk))
        token = default_token_generator.make_token(obj.kullanici)
        reset_url = request.build_absolute_uri(f'/portal/sifre-sifirla/{uid}/{token}/')
        
        # E-postayı gönder (konsola yazdırır)
        try:
            send_mail(
                'Site Yönetim Portalı - Şifre Sıfırlama',
                f'''
    Sayın {obj.kisi.ad_soyad},

    Şifrenizi sıfırlamak için aşağıdaki linke tıklayınız:

    {reset_url}

    Kullanıcı adınız: {obj.kullanici.username}

    Bu link 24 saat geçerlidir.

    İyi günler dileriz.
                ''',
                settings.DEFAULT_FROM_EMAIL,
                [obj.email],
                fail_silently=False
            )
            messages.success(request, f'✅ {obj.kisi.ad_soyad} adresine şifre sıfırlama linki gönderildi. (Konsolda görüntüleyin)')
        except Exception as e:
            messages.error(request, f'❌ E-posta gönderilemedi: {str(e)}')
        
        return redirect('admin:bina_dairekullanici_changelist')

admin_site.register(DaireKullanici, DaireKullaniciAdmin)

# Diğer modeller
admin_site.register(SiteAyarlari)
admin_site.register(Blok)
admin_site.register(Daire, DaireAdmin)
admin_site.register(Kisi, KisiAdmin)
admin_site.register(DaireIliskisi, DaireIliskisiAdmin)
admin_site.register(Aidat, AidatAdmin)
admin_site.register(Gider, GiderAdmin)
admin_site.register(Yonetici, YoneticiAdmin)
admin_site.register(Abonelik)
admin_site.register(Firma, FirmaAdmin)
admin_site.register(Banka, BankaAdmin)
admin_site.register(Depozito, DepozitoAdmin)
admin_site.register(DepozitoHareket, DepozitoHareketAdmin)
admin_site.register(Fatura)
admin_site.register(BankaHareket, BankaHareketAdmin)